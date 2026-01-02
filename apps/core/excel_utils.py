from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def xlsx_from_rows(
    rows: list[dict[str, Any]],
    *,
    sheet_name: str = "Sheet1",
    header_order: list[str] | None = None,
) -> bytes:
    """
    Create an XLSX file from a list of dict rows (keys -> columns).
    Returns bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Sheet1"

    if not rows:
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    # Determine headers
    keys = list(header_order) if header_order else list(rows[0].keys())
    # Ensure we include any extra keys from later rows
    if not header_order:
        seen = set(keys)
        for r in rows[1:]:
            for k in r.keys():
                if k not in seen:
                    keys.append(k)
                    seen.add(k)

    ws.append(keys)
    for r in rows:
        ws.append([r.get(k, "") for k in keys])

    # Simple column autosize
    for col_idx, key in enumerate(keys, start=1):
        max_len = len(str(key)) if key is not None else 0
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(10, max_len + 2))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def xlsx_to_rows(file_obj) -> list[dict[str, Any]]:
    """
    Read first worksheet of an Excel file and return list of dict rows.
    First row is treated as header.
    """
    wb = load_workbook(filename=file_obj, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header = [str(h).strip() if h is not None else "" for h in header]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        d = {}
        empty = True
        for i, col_name in enumerate(header):
            if not col_name:
                continue
            val = row[i] if i < len(row) else None
            if val is not None and str(val).strip() != "":
                empty = False
            d[col_name] = val
        if empty:
            continue
        out.append(d)
    return out




