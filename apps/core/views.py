from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, get_user_model, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from django.core.files.base import ContentFile
from django.conf import settings
from .models import SystemSetting, Tenant
# VisitTask modelini doğru adresten çağırıyoruz:
from apps.field_operations.models import VisitTask
from .utils import calculate_distance
from datetime import date
import json
from django.db.models import Count, Q as DQ
from django.http import HttpResponseForbidden
from django.http import HttpResponse
from apps.users.hierarchy_access import get_hierarchy_scope_for_user
from apps.users.utils import ensure_root_admin_configured, get_assigned_user_ids_under_admin_node, is_root_admin, get_root_admin_user
from apps.users.decorators import root_admin_required
from apps.users.models import UserRole, CustomUser


# Gerekli Modeller
from apps.field_operations.models import VisitTask
from apps.customers.models import Customer
from apps.forms.models import Survey, SurveyAnswer, Question

from .models import SystemSetting

def init_default_settings():
    """Sistemde hiç ayar yoksa varsayılanları oluşturur."""
    defaults = [
        # --- GENEL AYARLAR ---
        {
            'key': 'app_sync_interval',
            'label': 'Mobil Senkronizasyon Süresi (Dakika)',
            'value': '15',
            'category': 'general',
            'input_type': 'number',
            'description': 'Mobil uygulamanın sunucudan yeni verileri çekme sıklığı.'
        },
        {
            'key': 'maintenance_mode',
            'label': 'Bakım Modu',
            'value': 'False',
            'category': 'general',
            'input_type': 'bool',
            'description': 'Açılırsa sadece yöneticiler sisteme girebilir.'
        },
        
        # --- ZİYARET AYARLARI ---
        {
            'key': 'visit_radius',
            'label': 'Mağaza Giriş Mesafesi (Metre)',
            'value': '300', # 300 metre
            'category': 'visit',
            'input_type': 'number',
            'description': 'Personel mağazaya en fazla ne kadar uzaktayken ziyaret başlatabilir?'
        },
        {
            'key': 'distance_rule',
            'label': 'Mesafe Kuralı',
            'value': 'True',
            'category': 'visit',
            'input_type': 'bool',
            'description': 'Açık: Giriş mesafesi ve gezinme mesafesi kontrolü yapılır. Kapalı: Mesafe kontrolü yapılmaz, herhangi bir mesafeden ziyaret başlatılabilir.'
        },
        {
            'key': 'wander_radius',
            'label': 'Gezinme Sınırı (Metre)',
            'value': '500',
            'category': 'visit',
            'input_type': 'number',
            'description': 'Ziyaret sırasında mağaza konumundan maksimum uzaklaşma mesafesi. Bu mesafeyi aşarsa ziyaret otomatik bitirilir.'
        },
        
        # --- E-POSTA AYARLARI ---
        {
            'key': 'email_host',
            'label': 'SMTP Sunucu',
            'value': '',
            'category': 'email',
            'input_type': 'text',
            'description': 'SMTP sunucu adresi (örn: smtp.gmail.com, smtp.office365.com)'
        },
        {
            'key': 'email_port',
            'label': 'SMTP Port',
            'value': '587',
            'category': 'email',
            'input_type': 'number',
            'description': 'SMTP port numarası (genellikle 587 veya 465)'
        },
        {
            'key': 'email_use_tls',
            'label': 'TLS Kullan',
            'value': 'True',
            'category': 'email',
            'input_type': 'bool',
            'description': 'TLS şifreleme kullanılsın mı? (Genellikle açık olmalı)'
        },
        {
            'key': 'email_host_user',
            'label': 'SMTP Kullanıcı Adı',
            'value': '',
            'category': 'email',
            'input_type': 'text',
            'description': 'E-posta adresi (örn: noreply@yourdomain.com)'
        },
        {
            'key': 'email_host_password',
            'label': 'SMTP Şifresi',
            'value': '',
            'category': 'email',
            'input_type': 'password',
            'description': 'E-posta şifresi veya App Password (Gmail için)'
        },
        {
            'key': 'email_default_from_email',
            'label': 'Gönderen E-posta',
            'value': 'noreply@rotexia.com',
            'category': 'email',
            'input_type': 'text',
            'description': 'Gönderen olarak görünecek e-posta adresi'
        },
    ]

    for setting in defaults:
        # Eğer bu ayar veritabanında yoksa oluştur (tenant=None - Rotexia şablonu)
        if not SystemSetting.objects.filter(key=setting['key'], tenant__isnull=True).exists():
            SystemSetting.objects.create(**setting, tenant=None)

def sync_settings_to_all_tenants():
    """
    Rotexia'daki (tenant=None) ayar şablonlarını tüm firmalara kopyalar.
    Rotexia'da eklenen yeni ayarlar otomatik olarak tüm firmalara eklenir.
    """
    from .models import Tenant
    
    # Rotexia şablon ayarları (tenant=None)
    rotexia_settings = SystemSetting.objects.filter(tenant__isnull=True)
    
    # Tüm aktif firmalar
    tenants = Tenant.objects.filter(is_active=True)
    
    synced_count = 0
    for tenant in tenants:
        for template_setting in rotexia_settings:
            # Bu firma için bu ayar var mı kontrol et
            tenant_setting = SystemSetting.objects.filter(
                key=template_setting.key,
                tenant=tenant
            ).first()
            
            # Yoksa şablonundan kopyala
            if not tenant_setting:
                SystemSetting.objects.create(
                    tenant=tenant,
                    key=template_setting.key,
                    label=template_setting.label,
                    value=template_setting.value,  # Şablon değerini kopyala
                    category=template_setting.category,
                    input_type=template_setting.input_type,
                    description=template_setting.description
                )
                synced_count += 1
    
    return synced_count

# --- OTOMATİK GİRİŞ ---
def auto_login(request):
    User = get_user_model()
    user = User.objects.filter(is_superuser=True).first()
    if not user:
        user = User.objects.first()
    
    if user:
        if not user.is_active:
            user.is_active = True
            user.save()
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        return redirect('home')
    else:
        return render(request, 'base.html', {'content': 'Kullanıcı yok...'})
    

from .models import SystemSetting

@login_required
def settings_home(request):
    """
    Ayarlar sayfası - Admin paneli veya firma paneli
    Subdomain-only mod: request.tenant'a göre yönlendirir
    """
    # Middleware'den gelen tenant bilgisini kontrol et (subdomain-only mod)
    tenant = getattr(request, 'tenant', None)
    
    # Development modunda session'dan tenant al
    if not tenant:
        tenant_id = request.session.get('tenant_id') or request.session.get('connect_tenant_id')
        if tenant_id:
            try:
                tenant = Tenant.objects.get(id=tenant_id, is_active=True)
                request.tenant = tenant
            except Tenant.DoesNotExist:
                pass
    
    host = request.get_host()
    host_without_port = host.split(':')[0] if ':' in host else host
    
    # Subdomain kontrolü
    is_admin_subdomain = False
    if '.' in host_without_port and host_without_port not in ['127.0.0.1', 'localhost']:
        parts = host_without_port.split('.')
        subdomain = parts[0].lower()
        is_admin_subdomain = (subdomain == 'admin')
    
    # Admin subdomain'inde veya tenant yoksa -> admin ayarları
    admin_from_panel = request.session.get('admin_from_panel', False)
    if is_admin_subdomain or (not tenant and not admin_from_panel):
        if is_root_admin(request.user):
            return redirect('admin_settings')
        else:
            messages.error(request, 'Admin paneline erişim yetkiniz yok.')
            return redirect('home')
    
    # Firma subdomain'inde ve tenant var -> firma ayarları
    if tenant:
        return redirect('tenant_settings')
    
    # Fallback
    messages.warning(request, 'Ayarlar sayfasına erişilemedi.')
    return redirect('home')

@login_required
@root_admin_required
def admin_settings(request):
    """
    Admin Paneli Ayarları - Global/Şablon ayarlar
    Burada eklenen içerikler firmalara otomatik aktarılmaz
    Subdomain-only mod: Sadece admin.fieldops.com'dan erişilmeli
    """
    # Varsayılan ayarları oluştur (eğer yoksa)
    try:
        init_default_settings()
    except Exception as e:
        print(f"[WARNING] init_default_settings failed: {str(e)}")
    
    # Admin panelinde tenant olmamalı
    request.tenant = None
    
    # Subdomain kontrolü - admin subdomain'inde miyiz?
    host = request.get_host()
    host_without_port = host.split(':')[0] if ':' in host else host
    is_admin_subdomain = False
    if '.' in host_without_port:
        parts = host_without_port.split('.')
        subdomain = parts[0].lower()
        is_admin_subdomain = (subdomain == 'admin')
    
    # Admin subdomain'inde değilsek ve subdomain varsa, uyarı ver
    if not is_admin_subdomain and '.' in host_without_port:
        # Firma subdomain'indeyiz, admin ayarlarına erişim yok
        messages.warning(request, 'Admin ayarlarına sadece admin subdomain\'inden erişebilirsiniz.')
        # Firma subdomain'ine geri dön
        tenant = getattr(request, 'tenant', None)
        if tenant:
            if settings.DEBUG:
                return redirect(f"http://{tenant.slug}.localhost:8000/tenant/settings/")
            else:
                domain = getattr(settings, 'SUBDOMAIN_DOMAIN', 'fieldops.com')
                protocol = 'https' if not settings.DEBUG else 'http'
                return redirect(f"{protocol}://{tenant.slug}.{domain}/tenant/settings/")
    
    # --- POST İŞLEMLERİ ---
    if request.method == 'POST':
        # Handle role add/delete (tenant=None - admin paneli için global roller)
        if 'add_role' in request.POST:
            role_name = request.POST.get('role_name', '').strip()
            if role_name:
                # Admin panelinde tenant=None ile oluştur
                if not UserRole.objects.filter(name=role_name, tenant__isnull=True).exists():
                    UserRole.objects.create(name=role_name, tenant=None)
                    messages.success(request, f'✅ Rol "{role_name}" eklendi (Global Şablon).')
                else:
                    messages.warning(request, f'⚠️ Rol "{role_name}" zaten mevcut.')
            else:
                messages.error(request, '❌ Rol adı boş olamaz.')
        elif 'delete_role' in request.POST:
            role_id = request.POST.get('role_id')
            if role_id:
                try:
                    # Sadece tenant=None olan rolleri sil (admin paneli rolleri)
                    role = UserRole.objects.get(id=role_id, tenant__isnull=True)
                    role_name = role.name
                    role.delete()
                    messages.success(request, f'✅ Rol "{role_name}" silindi.')
                except UserRole.DoesNotExist:
                    messages.error(request, '❌ Rol bulunamadı veya bu rol silinemez.')
        
        # Handle regular settings update (global ayarlar - Rotexia şablon)
        all_settings = SystemSetting.objects.filter(tenant__isnull=True)
        for setting in all_settings:
            if setting.input_type == 'bool':
                new_val = 'True' if request.POST.get(setting.key) == 'on' else 'False'
            else:
                new_val = request.POST.get(setting.key)
            
            if new_val is not None:
                setting.value = new_val
                setting.save()
        
        # Rotexia'da ayar değiştiyse, tüm firmalara sync et
        if 'add_role' not in request.POST and 'delete_role' not in request.POST:
            sync_count = sync_settings_to_all_tenants()
            if sync_count > 0:
                messages.success(request, f'✅ Ayarlar kaydedildi ve {sync_count} yeni ayar tüm firmalara eklendi.')
            else:
                messages.success(request, '✅ Ayarlar kaydedildi.')
        return redirect('admin_settings')
    
    # --- VERİLERİ ÇEK (TENANT=None - GLOBAL/ROTEXIA ŞABLON) ---
    settings_general = SystemSetting.objects.filter(category='general', tenant__isnull=True)
    settings_visit = SystemSetting.objects.filter(category='visit', tenant__isnull=True)
    settings_user = SystemSetting.objects.filter(category='user', tenant__isnull=True)
    settings_email = SystemSetting.objects.filter(category='email', tenant__isnull=True).order_by('id')
    
    # Eğer mail ayarları yoksa, tekrar oluştur
    if not settings_email.exists():
        print("[DEBUG] Email settings not found, creating...")
        init_default_settings()
        settings_email = SystemSetting.objects.filter(category='email').order_by('id')
    
    # Admin panelinde sadece tenant=None olan rolleri göster (global şablonlar)
    user_roles = UserRole.objects.filter(tenant__isnull=True).order_by('name')
    
    root_admin = is_root_admin(request.user)
    print(f"[DEBUG] is_root_admin: {root_admin}, email settings count: {settings_email.count()}")

    context = {
        'is_admin_panel': True,
        'is_root_admin': root_admin,
        'settings_general': settings_general,
        'settings_visit': settings_visit,
        'settings_user': settings_user,
        'settings_email': settings_email,
        'user_roles': user_roles,
    }
    return render(request, 'apps/Core/settings.html', context)

@login_required
def tenant_settings(request):
    """
    Firma Paneli Ayarları - Tenant-specific ayarlar
    Sadece seçili firmanın kendi ayarları görünür
    """
    from apps.core.tenant_utils import get_current_tenant, filter_by_tenant
    from apps.core.models import Tenant
    
    # Önce request.tenant'ı kontrol et (middleware'den)
    tenant = getattr(request, 'tenant', None)
    
    # Eğer request.tenant yoksa, session'dan al (development modu için)
    if not tenant:
        tenant_id = request.session.get('tenant_id') or request.session.get('connect_tenant_id')
        if tenant_id:
            try:
                tenant = Tenant.objects.get(id=tenant_id, is_active=True)
                # Request'e de ekle (middleware için)
                request.tenant = tenant
            except Tenant.DoesNotExist:
                request.session.pop('tenant_id', None)
                request.session.pop('connect_tenant_id', None)
                tenant = None
    
    # Hala tenant yoksa, get_current_tenant'ı kullan (fallback)
    if not tenant:
        tenant = get_current_tenant(request)
        if tenant:
            request.tenant = tenant
    
    if not tenant:
        messages.error(request, '❌ Firma seçimi yapılmamış! Lütfen üst menüden bir firma seçin.')
        return redirect('home')
    
    # --- POST İŞLEMLERİ ---
    if request.method == 'POST':
        from apps.core.tenant_utils import set_tenant_on_save
        
        # Handle role add/delete (tenant-specific)
        if 'add_role' in request.POST:
            role_name = request.POST.get('role_name', '').strip()
            if role_name:
                # DEBUG: Tenant bilgisini kontrol et ve logla
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"DEBUG tenant_settings: tenant={tenant}, tenant_id={tenant.id if tenant else None}, tenant_name={tenant.name if tenant else None}")
                logger.error(f"DEBUG session: tenant_id={request.session.get('tenant_id', 'YOK')}")
                logger.error(f"DEBUG request.tenant: {getattr(request, 'tenant', 'YOK')}")
                
                if not tenant:
                    messages.error(request, '❌ Firma seçimi bulunamadı! Rol eklenemedi. Lütfen üst menüden bir firma seçin.')
                    return redirect('tenant_settings')
                
                # Bu firmanın rolleri arasında kontrol et
                existing_role = UserRole.objects.filter(name=role_name, tenant=tenant).first()
                if not existing_role:
                    # Tenant'ı direkt ata - MUTLAKA tenant_id ile
                    role = UserRole(name=role_name, tenant_id=tenant.id)  # tenant_id kullan
                    role.save()
                    
                    # DEBUG: Kayıt kontrolü - refresh from DB
                    role.refresh_from_db()
                    if role.tenant_id != tenant.id:
                        logger.error(f"DEBUG: Rol kaydedildi ama tenant yanlış! role.tenant_id={role.tenant_id}, tenant.id={tenant.id}")
                        messages.error(request, f'❌ Rol kaydedilirken hata oluştu! Tenant: {role.tenant_id} != {tenant.id}')
                    else:
                        logger.error(f"DEBUG: Rol başarıyla kaydedildi! role.tenant_id={role.tenant_id}, tenant.id={tenant.id}")
                        messages.success(request, f'✅ Rol "{role_name}" "{tenant.name}" firmasına eklendi.')
                else:
                    messages.warning(request, f'⚠️ Rol "{role_name}" bu firmada zaten mevcut.')
            else:
                messages.error(request, '❌ Rol adı boş olamaz.')
        elif 'delete_role' in request.POST:
            role_id = request.POST.get('role_id')
            if role_id:
                try:
                    # Sadece bu firmanın rollerini sil
                    role = get_object_or_404(filter_by_tenant(UserRole.objects.all(), request), id=role_id)
                    role_name = role.name
                    role.delete()
                    messages.success(request, f'✅ Rol "{role_name}" silindi.')
                except:
                    messages.error(request, '❌ Rol bulunamadı veya bu rol silinemez.')
        
        # Handle regular settings update (tenant-specific)
        # Önce Rotexia şablon ayarlarını bu firmaya sync et (yoksa ekle)
        sync_settings_to_all_tenants()
        
        # Şimdi tenant'a özel ayarları güncelle
        all_settings = SystemSetting.objects.filter(tenant=tenant)
        updated_count = 0
        for setting in all_settings:
            if setting.input_type == 'bool':
                new_val = 'True' if request.POST.get(setting.key) == 'on' else 'False'
            else:
                new_val = request.POST.get(setting.key, '').strip()  # Boş string'i de al
            
            # Boş string veya None kontrolü (password için özel durum)
            if setting.input_type == 'password' and not new_val:
                # Password alanı boşsa, mevcut değeri koru (değiştirme)
                continue
            
            # Değer varsa güncelle
            if new_val is not None and new_val != '':
                setting.value = new_val
                setting.save(update_fields=['value'])
                updated_count += 1
        
        if 'add_role' not in request.POST and 'delete_role' not in request.POST:
            if updated_count > 0:
                messages.success(request, f'✅ {updated_count} ayar kaydedildi.')
            else:
                messages.info(request, 'ℹ️ Değişiklik yapılmadı.')
        return redirect('tenant_settings')
    
    # Rotexia şablon ayarlarını bu firmaya sync et (yoksa ekle)
    sync_settings_to_all_tenants()
    
    # --- VERİLERİ ÇEK (TENANT-SPECIFIC) ---
    settings_general = SystemSetting.objects.filter(tenant=tenant, category='general')
    settings_visit = SystemSetting.objects.filter(tenant=tenant, category='visit')
    settings_user = SystemSetting.objects.filter(tenant=tenant, category='user')
    settings_email = SystemSetting.objects.filter(tenant=tenant, category='email').order_by('id')
    
    # Sadece bu firmanın rolleri
    user_roles = filter_by_tenant(UserRole.objects.all(), request).order_by('name')

    context = {
        'is_admin_panel': False,
        'tenant': tenant,
        'settings_general': settings_general,
        'settings_visit': settings_visit,
        'settings_user': settings_user,
        'settings_email': settings_email,
        'user_roles': user_roles,
    }
    return render(request, 'apps/Core/settings.html', context)

# Eğer eski require_gps ayarı varsa, distance_rule olarak güncelle (migration helper)
def migrate_old_settings():
    old_require_gps = SystemSetting.objects.filter(key='require_gps').first()
    if old_require_gps:
        # Yeni ayar zaten var mı kontrol et
        if not SystemSetting.objects.filter(key='distance_rule').exists():
            # Eski ayarı yeni isimle güncelle
            old_require_gps.key = 'distance_rule'
            old_require_gps.label = 'Mesafe Kuralı'
            old_require_gps.description = 'Açık: Giriş mesafesi ve gezinme mesafesi kontrolü yapılır. Kapalı: Mesafe kontrolü yapılmaz, herhangi bir mesafeden ziyaret başlatılabilir.'
            old_require_gps.save()
        else:
            # Yeni ayar zaten varsa, eski ayarı sil
            old_require_gps.delete()
    return None

# --- AKILLI ANASAYFA ---
def index(request):
    """Ana sayfa - Firma adı girme ekranı (PUBLIC - Herkes erişebilir)"""
    # Eğer kullanıcı zaten giriş yapmışsa, root admin ise admin_home'a, değilse logout yap
    if request.user.is_authenticated:
        from apps.users.utils import is_root_admin
        if is_root_admin(request.user):
            return redirect('admin_home')
        else:
            # Normal kullanıcı için subdomain gerekli - logout yapıp ana sayfayı göster
            from django.contrib.auth import logout
            logout(request)
    
    return render(request, 'index.html')

def _is_mobile_device(request):
    """Mobil cihaz kontrolü"""
    # Development'ta test için query parametresi kontrolü
    if request.GET.get('mobile') == '1':
        return True
    
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    mobile_keywords = ['mobile', 'android', 'iphone', 'ipad', 'webos', 'ipod']
    return any(keyword in user_agent for keyword in mobile_keywords)

def company_connect(request):
    """Firma adı girildikten sonra, login sayfasına yönlendir (dropdown ile seçim yapılacak)"""
    if request.method == 'POST':
        company_name = request.POST.get('company_name', '').strip()
        
        if not company_name:
            messages.error(request, 'Lütfen firma adı girin.')
            return redirect('index')
        
        # Mobil cihaz kontrolü
        is_mobile = _is_mobile_device(request)
        
        # "Rotexia" (ana sistem) yazıldıysa: Root admin giriş ekranına yönlendir.
        # KURAL: Firma adıyla bağlanıldıysa, o firmadan başka bir sisteme geçiş OLMAZ.
        if company_name.lower() in ['rotexia', 'sistem', 'admin']:
            # Tenant session'ını temizle: Rotexia akışı tenant seçimi değildir.
            for key in ['tenant_id', 'connect_tenant_id', 'connect_tenant_slug', 'connect_tenant_color', 'connect_tenant_name', 'admin_from_panel']:
                request.session.pop(key, None)

            # Mobil ise mobil login'e, değilse admin login'e yönlendir.
            if is_mobile:
                return redirect('mobile_login')
            return redirect('admin_login')
        
        # Eski mantık: Firma adından tenant'ı bul (geriye dönük uyumluluk için)
        from django.utils.text import slugify
        slug = slugify(company_name)
        
        try:
            tenant = Tenant.objects.filter(
                models.Q(slug=slug) | models.Q(name__iexact=company_name),
                is_active=True
            ).first()
            
            if not tenant:
                messages.error(request, f'Böyle bir firma bulunmamaktadır. Lütfen firma adını kontrol edin.')
                return redirect('index')
            
            # Firma bilgilerini session'a kaydet (login sayfasında kullanılacak)
            request.session['connect_tenant_id'] = tenant.id
            request.session['connect_tenant_slug'] = tenant.slug
            request.session['connect_tenant_color'] = tenant.primary_color
            request.session['connect_tenant_name'] = tenant.name
            
            # Mobil ise mobil login'e, değilse normal login'e yönlendir
            if is_mobile:
                return redirect('mobile_login')
            return redirect('login')
            
        except Exception as e:
            messages.error(request, f'Firma bulunurken hata oluştu: {str(e)}')
            return redirect('index')
    
    return redirect('index')

def login_with_tenant(request, tenant_slug):
    """Firma rengine göre dinamik login sayfası - Legacy support, ana sayfaya yönlendir"""
    # Bu URL artık kullanılmıyor, önce ana sayfaya gitmesi gerekiyor
    try:
        tenant = Tenant.objects.get(slug=tenant_slug, is_active=True)
        # Session'a tenant bilgilerini kaydet
        request.session['connect_tenant_id'] = tenant.id
        request.session['connect_tenant_slug'] = tenant.slug
        request.session['connect_tenant_color'] = tenant.primary_color
        request.session['connect_tenant_name'] = tenant.name
        # Login sayfasına yönlendir
        return redirect('login')
    except Tenant.DoesNotExist:
        messages.error(request, 'Firma bulunamadı.')
        return redirect('index')

def home(request):
    """
    Dashboard sayfası - Development'ta session bazlı, production'da subdomain bazlı
    
    Development: localhost:8000 -> Session'dan tenant bilgisi alınır
    Production: firma1.fieldops.com -> Middleware'den tenant bilgisi gelir
    """
    # MANUEL GİRİŞ KONTROLÜ - Her zaman giriş yapılması gerekiyor
    if not request.user.is_authenticated:
        messages.error(request, 'Lütfen giriş yapın.')
        return redirect('index')
    
    # Önce middleware'den gelen tenant bilgisini kontrol et
    tenant = getattr(request, 'tenant', None)
    host = request.get_host()
    host_without_port = host.split(':')[0] if ':' in host else host
    
    # Render veya benzeri servislerde subdomain yok, session bazlı çalış
    # Render domain'leri: *.onrender.com
    is_render_domain = 'onrender.com' in host_without_port
    
    # Subdomain kontrolü (Render ve localhost hariç)
    has_subdomain = (
        '.' in host_without_port and 
        host_without_port not in ['127.0.0.1', 'localhost'] and
        not is_render_domain  # Render'da subdomain yok, session bazlı
    )
    subdomain = None
    if has_subdomain:
        parts = host_without_port.split('.')
        subdomain = parts[0].lower()
    
    # Tenant'ı middleware'den al (subdomain veya session'dan)
    tenant = getattr(request, 'tenant', None)
    
    # Admin panelinden bağlanıldıysa (admin_from_panel), session'dan tenant'ı oku
    admin_from_panel = request.session.get('admin_from_panel', False)
    if admin_from_panel and not tenant:
        tenant_id = request.session.get('tenant_id') or request.session.get('connect_tenant_id')
        if tenant_id:
            try:
                tenant = Tenant.objects.get(id=tenant_id, is_active=True)
                request.tenant = tenant  # Request'e ekle
            except Tenant.DoesNotExist:
                # Geçersiz tenant_id, session'dan temizle
                request.session.pop('tenant_id', None)
                request.session.pop('connect_tenant_id', None)
                request.session.pop('admin_from_panel', None)
    
    # Subdomain yoksa (development - localhost:8000 veya 127.0.0.1:8000)
    if not has_subdomain:
        # Middleware zaten session'dan tenant'ı okumuş olmalı
        # Ama eğer okumadıysa, session'dan tekrar oku
        if not tenant:
            tenant_id = request.session.get('tenant_id') or request.session.get('connect_tenant_id')
            if tenant_id:
                try:
                    tenant = Tenant.objects.get(id=tenant_id, is_active=True)
                    request.tenant = tenant  # Request'e ekle
                except Tenant.DoesNotExist:
                    # Geçersiz tenant_id, session'dan temizle
                    request.session.pop('tenant_id', None)
                    request.session.pop('connect_tenant_id', None)
            
            # Hala tenant yoksa ve root admin değilse ana sayfaya yönlendir
            if not tenant:
                # Root admin ise ve admin_from_panel yoksa admin paneline yönlendir
                if is_root_admin(request.user) and not admin_from_panel:
                    messages.info(request, 'Lütfen bir firmayı seçin veya admin panelinden işlem yapın.')
                    return redirect('admin_home')
                messages.error(request, 'Firma bilgisi bulunamadı. Lütfen firma adınızı girin.')
                return redirect('index')
    
    # Subdomain var ama tenant bulunamadı
    if has_subdomain and not tenant:
        messages.error(request, f'"{subdomain}" subdomain\'i için firma bulunamadı. Lütfen admin panelinden firma oluşturun.')
        if is_root_admin(request.user):
            return redirect('admin_home')
        else:
            return redirect('index')
    
    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    mobile_keywords = ['mobile', 'android', 'iphone', 'ipad', 'webos', 'ipod']
    
    is_mobile = any(keyword in user_agent for keyword in mobile_keywords)
    
    if is_mobile:
        return redirect('mobile_home')
    else:
        # MASAÜSTÜ DASHBOARD - İSTEK ÜZERİNE BOŞALTILDI
        # Eski KPI hesaplamaları kaldırıldı.
        context = {
            'tenant': tenant,
            'kpi': {} 
        }
        return render(request, 'apps/Core/home.html', context)


def healthz(request):
    """
    Render/healthcheck endpoint. Always returns 200.
    """
    return HttpResponse("ok", content_type="text/plain")

# --- ÖZEL GİRİŞ VIEW (Firma Adı ile) ---
from django.contrib.auth.views import LoginView
from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_protect
from django.utils.decorators import method_decorator
from django.urls import reverse_lazy
from django.utils.text import slugify

@method_decorator(csrf_protect, name='dispatch')
class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    redirect_authenticated_user = True
    
    def get(self, request, *args, **kwargs):
        """GET request - Eğer tenant bilgisi yoksa ana sayfaya yönlendir"""
        # Session'da tenant bilgisi var mı kontrol et
        tenant_id = request.session.get('connect_tenant_id')
        tenant_slug = request.session.get('connect_tenant_slug')
        
        # Eğer tenant bilgisi yoksa, ana sayfaya (Bağlan sayfasına) yönlendir
        if not tenant_id or not tenant_slug:
            return redirect('index')
        
        # Tenant bilgisi varsa login_tenant.html template'ini kullan
        try:
            tenant = Tenant.objects.get(id=tenant_id, slug=tenant_slug, is_active=True)
            context = {
                'tenant': tenant,
                'primary_color': request.session.get('connect_tenant_color', tenant.primary_color),
            }
            return render(request, 'registration/login_tenant.html', context)
        except Tenant.DoesNotExist:
            # Tenant bulunamazsa session'ı temizle ve ana sayfaya yönlendir
            request.session.pop('connect_tenant_id', None)
            request.session.pop('connect_tenant_slug', None)
            request.session.pop('connect_tenant_color', None)
            request.session.pop('connect_tenant_name', None)
            return redirect('index')
    
    def post(self, request, *args, **kwargs):
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # Tenant bilgisini session'dan al
        tenant_id = request.session.get('connect_tenant_id')
        tenant_slug = request.session.get('connect_tenant_slug')
        
        # Eğer session'da tenant bilgisi yoksa ana sayfaya yönlendir
        if not tenant_id or not tenant_slug:
            messages.error(request, 'Lütfen önce firma adını girin.')
            return redirect('index')
        
        # Tenant'ı bul
        try:
            tenant = Tenant.objects.get(id=tenant_id, slug=tenant_slug, is_active=True)
        except Tenant.DoesNotExist:
            messages.error(request, 'Firma bulunamadı. Lütfen tekrar deneyin.')
            request.session.pop('connect_tenant_id', None)
            request.session.pop('connect_tenant_slug', None)
            return redirect('index')
        
        # Kullanıcıyı doğrula
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            # Root admin kontrolü - root admin her tenant'a giriş yapabilir
            from apps.users.utils import is_root_admin
            
            if not is_root_admin(user):
                # Normal kullanıcı için tenant kontrolü
                # Kullanıcının tenant'ı varsa ve seçilen tenant ile eşleşmiyorsa hata ver
                if hasattr(user, 'tenant') and user.tenant:
                    if user.tenant != tenant:
                        messages.error(request, f'Bu kullanıcı "{user.tenant.name}" firmasına aittir. Lütfen doğru firmayı seçin.')
                        return self.form_invalid(self.get_form())
                else:
                    # Kullanıcının tenant'ı yoksa, seçilen tenant'ı ata (eski kullanıcılar için)
                    user.tenant = tenant
                    user.save()
                    messages.info(request, f'Kullanıcınız "{tenant.name}" firmasına atandı.')
            
            login(request, user)
            # Tenant'ı session'a kaydet
            request.session['tenant_id'] = tenant.id
            # Bağlanma session bilgilerini temizle (artık gerekli değil)
            request.session.pop('connect_tenant_id', None)
            request.session.pop('connect_tenant_slug', None)
            request.session.pop('connect_tenant_color', None)
            request.session.pop('connect_tenant_name', None)
            # Redirect URL'i belirle
            redirect_url = self.get_success_url()
            return redirect(redirect_url)
        else:
            messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
            return self.form_invalid(self.get_form())
    

# --- GELİŞTİRİCİ ADMIN GİRİŞİ ---
@method_decorator(csrf_protect, name='dispatch')
class AdminLoginView(LoginView):
    template_name = 'registration/admin_login.html'
    redirect_authenticated_user = True
    
    def post(self, request, *args, **kwargs):
        """Admin girişi için firma adı gerektirmez, sadece kullanıcı adı ve şifre"""
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # Kullanıcıyı doğrula (firma kontrolü yok)
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            login(request, user)
            # Admin girişinde tenant session'ını temizle
            if 'tenant_id' in request.session:
                del request.session['tenant_id']
            # Redirect URL'i belirle
            redirect_url = self.get_success_url()
            return redirect(redirect_url)
        else:
            messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
            return self.form_invalid(self.get_form())
    
    def get_success_url(self):
        # Root admin kontrolü
        if is_root_admin(self.request.user):
            return reverse_lazy('admin_home')
        return reverse_lazy('home')

# --- LOGOUT VIEW ---
from django.contrib.auth import logout as auth_logout

def logout_view(request):
    """Logout - Tenant varsa o tenant'ın login sayfasına yönlendir, root admin ise admin login'e"""
    # Mobil cihaz kontrolü
    is_mobile = _is_mobile_device(request)
    
    # Önce kullanıcının root admin olup olmadığını kontrol et (logout'tan önce)
    is_root_admin_user = False
    if request.user.is_authenticated:
        try:
            from apps.users.utils import is_root_admin
            is_root_admin_user = is_root_admin(request.user)
        except:
            pass
    
    # Önce tenant bilgilerini al (logout'tan önce - session'dan)
    tenant_id = request.session.get('tenant_id') or request.session.get('connect_tenant_id')
    tenant_slug = request.session.get('connect_tenant_slug')
    tenant_name = request.session.get('connect_tenant_name')
    tenant_color = request.session.get('connect_tenant_color')
    
    # Kullanıcının tenant'ını kontrol et (user objesinden)
    user_tenant_id = None
    user_tenant_slug = None
    user_tenant_name = None
    user_tenant_color = None
    
    if request.user.is_authenticated and hasattr(request.user, 'tenant') and request.user.tenant:
        user_tenant_id = request.user.tenant.id
        user_tenant_slug = request.user.tenant.slug
        user_tenant_name = request.user.tenant.name
        user_tenant_color = request.user.tenant.primary_color
    
    # Logout yap (kullanıcı oturumunu sonlandır)
    auth_logout(request)
    
    # Root admin ise admin login sayfasına yönlendir
    if is_root_admin_user:
        # Tüm session'ı temizle
        for key in ['tenant_id', 'connect_tenant_id', 'connect_tenant_slug', 'connect_tenant_color', 'connect_tenant_name', 'admin_from_panel', 'from_tenant_logout', 'logout_tenant_slug', 'logout_tenant_name']:
            request.session.pop(key, None)
        # Mobil ise mobil login'e, değilse normal login'e yönlendir
        if is_mobile:
            return redirect('mobile_login')
        return redirect('admin_login')
    
    # Tenant kullanıcısı çıkışı
    # Logout'tan önce hangi tenant bilgisini kullanacağımızı belirle
    # Öncelik user'ın tenant'ına verilir, yoksa session'dakine
    final_tenant_id = user_tenant_id or tenant_id
    final_tenant_slug = user_tenant_slug or tenant_slug
    final_tenant_name = user_tenant_name or tenant_name
    final_tenant_color = user_tenant_color or tenant_color
    
    # Eğer tenant varsa, tenant bilgilerini session'a tekrar kaydet (logout sonrası)
    if final_tenant_id and final_tenant_slug:
        # Tenant bilgilerini session'a kaydet (login sayfasında kullanılacak)
        request.session['connect_tenant_id'] = final_tenant_id
        request.session['connect_tenant_slug'] = final_tenant_slug
        request.session['connect_tenant_name'] = final_tenant_name
        if final_tenant_color:
            request.session['connect_tenant_color'] = final_tenant_color
        
        # Logout flag'i ekle (login sayfasında mesaj göstermek için)
        request.session['from_tenant_logout'] = True
        request.session['logout_tenant_name'] = final_tenant_name or 'Firma'
        
        # Temizle: aktif tenant ID'si ve admin panel flag'i
        request.session.pop('tenant_id', None)
        request.session.pop('admin_from_panel', None)
        
        # Mobil ise mobil login'e, değilse normal login'e yönlendir
        if is_mobile:
            return redirect('mobile_login')
        return redirect('login')
    else:
        # Tenant yoksa tüm session'ı temizle
        for key in ['tenant_id', 'connect_tenant_id', 'connect_tenant_slug', 'connect_tenant_color', 'connect_tenant_name', 'admin_from_panel', 'from_tenant_logout', 'logout_tenant_slug', 'logout_tenant_name']:
            request.session.pop(key, None)
        return redirect('index')  # Ana sayfaya yönlendir

# --- ADMIN AYARLARI GÜNCELLEME ---
@login_required
@root_admin_required
def admin_update_settings(request):
    """Admin kullanıcı adı ve şifre güncelleme"""
    if request.method == 'POST':
        new_username = request.POST.get('username', '').strip()
        new_password = request.POST.get('password', '').strip()
        password_confirm = request.POST.get('password_confirm', '').strip()
        
        if not new_username:
            messages.error(request, 'Kullanıcı adı boş olamaz.')
            return redirect(request.META.get('HTTP_REFERER', 'admin_home'))
        
        # Kullanıcı adı değiştir
        if new_username != request.user.username:
            # Kullanıcı adı benzersiz mi kontrol et
            from django.contrib.auth import get_user_model
            User = get_user_model()
            if User.objects.filter(username=new_username).exclude(id=request.user.id).exists():
                messages.error(request, 'Bu kullanıcı adı zaten kullanılıyor.')
                return redirect(request.META.get('HTTP_REFERER', 'admin_home'))
            request.user.username = new_username
        
        # Şifre değiştir
        if new_password:
            if new_password != password_confirm:
                messages.error(request, 'Şifreler eşleşmiyor.')
                return redirect(request.META.get('HTTP_REFERER', 'admin_home'))
            if len(new_password) < 6:
                messages.error(request, 'Şifre en az 6 karakter olmalıdır.')
                return redirect(request.META.get('HTTP_REFERER', 'admin_home'))
            request.user.set_password(new_password)
        
        request.user.save()
        messages.success(request, 'Ayarlar başarıyla güncellendi. Yeniden giriş yapmanız gerekiyor.')
        
        # Çıkış yap ve login sayfasına yönlendir
        from django.contrib.auth import logout
        logout(request)
        return redirect('login')
    
    return redirect('admin_home')

# --- GELİŞTİRİCİ ADMIN ANA SAYFA (Firma Listesi) ---
from django.views.decorators.cache import never_cache

@never_cache
@login_required
@root_admin_required
def admin_home(request):
    """
    Root admin için firma listesi ve yönetim sayfası
    Subdomain-only mod: Bu view sadece admin.fieldops.com'dan erişilmeli
    """
    # Admin panelinde tenant olmamalı - Session'daki tenant bilgilerini temizle
    request.tenant = None
    
    # Session temizliği: Admin ana sayfasına gelindiyse, herhangi bir tenant bağlantısı kesilmeli
    # admin_from_panel dahil her şeyi temizle - HER ZAMAN YAPILIR
    keys_to_clear = [
        'tenant_id', 
        'connect_tenant_id', 
        'connect_tenant_slug', 
        'connect_tenant_color', 
        'connect_tenant_name',
        'admin_from_panel'
    ]
    for key in keys_to_clear:
        request.session.pop(key, None)
    
    # Session değişikliklerini kaydet
    request.session.modified = True
    
    tenants = Tenant.objects.filter(is_active=True).order_by('-created_at')
    
    # Eksik admin kullanıcılarını kontrol et
    User = get_user_model()
    tenants_without_admin = []
    for tenant in tenants:
        admin_user = User.objects.filter(
            tenant=tenant,
            user_code='admin',
            authority='Admin'
        ).first()
        if not admin_user:
            tenants_without_admin.append(tenant)
    
    context = {
        'tenants': tenants,
        'tenants_without_admin': tenants_without_admin,
        'is_admin_panel': True,
        'tenant': None,
    }
    return render(request, 'apps/Core/admin_home.html', context)

# --- FİRMA DÜZENLEME ---
@login_required
@root_admin_required
def edit_tenant(request, tenant_id):
    """Firma düzenleme sayfası"""
    tenant = get_object_or_404(Tenant, id=tenant_id)
    
    # Menü ayarları için varsayılan değerler (ana menüler ve alt menüler)
    default_menus = {
        'hierarchy': True,
        'users': True,
        'customers': True,
        'tasks': True,
        'route_plan': True,
        'forms': True,
        'images': True,
        'reports': True,
        # Alt menüler
        'reports_list': True,  # Raporlar listesi
        'reports_trash': True,  # Çöp Kutusu
    }
    
    # Mevcut menu_settings yoksa varsayılan değerleri kullan
    if not tenant.menu_settings or len(tenant.menu_settings) == 0:
        tenant.menu_settings = default_menus
        tenant.save(update_fields=['menu_settings'])
    
    # Mevcut menu_settings'te eksik olan menüleri ekle (yeni menüler için)
    menu_settings_updated = tenant.menu_settings.copy()
    for key, default_value in default_menus.items():
        if key not in menu_settings_updated:
            menu_settings_updated[key] = default_value
    
    if menu_settings_updated != tenant.menu_settings:
        tenant.menu_settings = menu_settings_updated
        tenant.save(update_fields=['menu_settings'])
    
    if request.method == 'POST':
        # İsim güncelleme
        if 'name' in request.POST:
            tenant.name = request.POST.get('name', '').strip()
        
        # Renk güncelleme
        if 'primary_color' in request.POST:
            tenant.primary_color = request.POST.get('primary_color', '#0d6efd')
        
        # Logo güncelleme
        if 'logo' in request.FILES:
            tenant.logo = request.FILES['logo']
        
        # Superuser bilgileri güncelleme
        if 'superuser_username' in request.POST:
            new_username = request.POST.get('superuser_username', '').strip()
            new_password = request.POST.get('superuser_password', '').strip()
            
            # Kullanıcı adı değiştiyse
            if new_username != tenant.superuser_username:
                # Eski kullanıcıyı bul ve username'i güncelle
                User = get_user_model()
                old_user = User.objects.filter(username=tenant.superuser_username, tenant=tenant).first()
                if old_user:
                    old_user.username = new_username
                    old_user.save()
                
                tenant.superuser_username = new_username
            
            # Şifre girildiyse güncelle
            if new_password:
                tenant.superuser_plain_password = new_password
                # Kullanıcının şifresini güncelle
                User = get_user_model()
                user = User.objects.filter(username=tenant.superuser_username, tenant=tenant).first()
                if user:
                    user.set_password(new_password)
                    user.save()
        
        # Menü ayarları güncelleme
        menu_settings = {}
        for menu_key in default_menus.keys():
            menu_settings[menu_key] = request.POST.get(f'menu_{menu_key}', 'off') == 'on'
        
        # Alt menü kontrolü: Ana menü kapalıysa alt menüler de kapalı olmalı
        if not menu_settings.get('reports', True):
            menu_settings['reports_list'] = False
            menu_settings['reports_trash'] = False
        
        tenant.menu_settings = menu_settings
        
        # Tenant'ı refresh etmeden önce kaydet
        tenant.save()
        
        # Tenant'ı veritabanından yeniden yükle (cache temizleme için)
        tenant.refresh_from_db()
        
        messages.success(request, f'✅ "{tenant.name}" firması ve superuser bilgileri güncellendi.')
        return redirect('admin_home')
    
    # Admin panelinde sidebar rengi değişmemesi için tenant'ı None yap
    # Ama formda tenant bilgisini gösteriyoruz, o yüzden ayrı değişkende tutalım
    sidebar_tenant = None  # Admin panelinde sidebar için tenant yok
    request.tenant = None  # Context processor için
    
    # Menü yapısı (ana menüler ve alt menüler)
    menu_structure = [
        {
            'key': 'hierarchy',
            'label': 'Hiyerarşi',
            'icon': 'fa-sitemap',
            'children': []
        },
        {
            'key': 'users',
            'label': 'Kullanıcılar',
            'icon': 'fa-users',
            'children': []
        },
        {
            'key': 'customers',
            'label': 'Müşteriler',
            'icon': 'fa-store',
            'children': []
        },
        {
            'key': 'tasks',
            'label': 'Görevler',
            'icon': 'fa-tasks',
            'children': []
        },
        {
            'key': 'route_plan',
            'label': 'Rota Planı',
            'icon': 'fa-route',
            'children': []
        },
        {
            'key': 'forms',
            'label': 'Formlar',
            'icon': 'fa-clipboard-list',
            'children': []
        },
        {
            'key': 'images',
            'label': 'Görseller',
            'icon': 'fa-images',
            'children': []
        },
        {
            'key': 'reports',
            'label': 'Raporlar',
            'icon': 'fa-chart-line',
            'children': [
                {'key': 'reports_list', 'label': 'Raporlar Listesi', 'icon': 'fa-list'},
                {'key': 'reports_trash', 'label': 'Çöp Kutusu', 'icon': 'fa-trash'},
            ]
        },
    ]
    
    context = {
        'tenant': tenant,  # Form için tenant bilgisi
        'sidebar_tenant': sidebar_tenant,  # Sidebar için None
        'is_admin_panel': True,  # Admin panelinde olduğumuzu belirt
        'menu_structure': menu_structure,
    }
    return render(request, 'apps/Core/edit_tenant.html', context)

# --- FİRMA EKLEME ---
@login_required
def create_company(request):
    """Yeni firma oluştur - Subdomain otomatik oluşturulur"""
    if not is_root_admin(request.user):
        return HttpResponseForbidden("Bu işlem için yetkiniz yok.")
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        primary_color = request.POST.get('primary_color', '#0d6efd')
        # Slug manuel girilebilir veya otomatik oluşturulur
        slug_input = request.POST.get('slug', '').strip()
        
        if not name:
            messages.error(request, 'Firma adı gereklidir.')
            return redirect('admin_home')
        
        try:
            # Slug oluştur (manuel girilmişse onu kullan, yoksa otomatik oluştur)
            if slug_input:
                slug = slugify(slug_input)
            else:
                slug = slugify(name)
            
            # Slug'un boş olmaması için kontrol
            if not slug:
                messages.error(request, 'Firma adından geçerli bir subdomain oluşturulamadı. Lütfen manuel olarak girin.')
                return redirect('admin_home')
            
            # Slug benzersiz olmalı - eğer varsa numara ekle
            original_slug = slug
            counter = 1
            while Tenant.objects.filter(slug=slug).exists():
                slug = f"{original_slug}-{counter}"
                counter += 1
                if counter > 100:  # Güvenlik için limit
                    messages.error(request, 'Çok fazla benzer firma adı var. Lütfen farklı bir ad seçin.')
                    return redirect('admin_home')
            
            # Varsayılan plan oluştur veya al
            from .models import Plan
            default_plan, _ = Plan.objects.get_or_create(
                name="Ücretsiz Plan",
                defaults={
                    'plan_type': 'basic',
                    'price_monthly': 0,
                    'max_users': 3,
                    'max_customers': 20,
                    'max_tasks_per_month': 100,
                }
            )
            
            # Superuser kullanıcı adı ve şifre (varsayılan)
            superuser_username = request.POST.get('superuser_username', f"{slug}_admin").strip()
            superuser_password = request.POST.get('superuser_password', 'admin123').strip()
            
            tenant = Tenant.objects.create(
                name=name,
                slug=slug,
                email=email or 'info@example.com',
                plan=default_plan,
                primary_color=primary_color,
                is_active=True,
                superuser_username=superuser_username,
                superuser_plain_password=superuser_password
            )
            
            # Firma için superuser oluştur
            User = get_user_model()
            
            # Tenant'ta kayıtlı kullanıcı adı ve şifreyi kullan
            admin_username = tenant.superuser_username
            admin_password = tenant.superuser_plain_password
            
            # Eğer bu username zaten varsa, farklı bir username kullan
            counter = 1
            original_admin_username = admin_username
            while User.objects.filter(username=admin_username).exists():
                admin_username = f"{original_admin_username}_{counter}"
                counter += 1
            
            # Kullanıcı adı değiştiyse tenant'ı güncelle
            if admin_username != tenant.superuser_username:
                tenant.superuser_username = admin_username
                tenant.save(update_fields=['superuser_username'])
            
            # Superuser oluştur
            admin_user = User.objects.create(
                username=admin_username,
                user_code='admin',
                first_name='Admin',
                last_name=name,
                email=email or f'admin@{slug}.fieldops.com',
                tenant=tenant,
                authority='Admin',
                is_staff=True,
                is_superuser=True,
                is_active=True
            )
            
            # Şifreyi set et
            admin_user.set_password(admin_password)
            admin_user.save()
            
            # Admin rolü oluştur (eğer yoksa)
            admin_role, _ = UserRole.objects.get_or_create(
                name='Admin',
                tenant=tenant,
                defaults={'description': 'Firma yöneticisi'}
            )
            admin_user.role = admin_role
            admin_user.save(update_fields=['role'])
            
            # Başarı mesajında subdomain bilgisi de göster
            messages.success(
                request, 
                f'✅ Firma "{name}" başarıyla oluşturuldu! '
                f'Subdomain: <strong>{slug}.fieldops.com</strong>'
            )
        except Exception as e:
            messages.error(request, f'❌ Hata: {str(e)}')
    
    return redirect('admin_home')

# --- EKSİK ADMIN KULLANICILARINI OLUŞTUR ---
@login_required
@root_admin_required
def login_as_tenant_superuser(request, tenant_id):
    """Root admin firmanın superuser'ı olarak giriş yapar"""
    if not is_root_admin(request.user):
        return HttpResponseForbidden("Bu işlem için yetkiniz yok.")
    
    tenant = get_object_or_404(Tenant, id=tenant_id)
    
    # Tenant'ın superuser kullanıcısını bul
    User = get_user_model()
    
    # Önce tenant'ta kayıtlı superuser_username ile ara
    superuser = None
    if tenant.superuser_username:
        superuser = User.objects.filter(username=tenant.superuser_username, tenant=tenant, is_superuser=True).first()
    
    # Bulunamazsa, tenant'ın admin kullanıcısını ara
    if not superuser:
        superuser = User.objects.filter(
            tenant=tenant,
            authority='Admin',
            is_superuser=True
        ).first()
    
    # Hala bulunamazsa, user_code='admin' ile ara
    if not superuser:
        superuser = User.objects.filter(
            tenant=tenant,
            user_code='admin',
            is_superuser=True
        ).first()
    
    if not superuser:
        messages.error(
            request, 
            f'❌ "{tenant.name}" firması için superuser bulunamadı. '
            f'Lütfen firma düzenleme sayfasından superuser bilgilerini kontrol edin.'
        )
        return redirect('admin_home')
    
    # Session'a tenant bilgilerini kaydet
    request.session['connect_tenant_id'] = tenant.id
    request.session['connect_tenant_slug'] = tenant.slug
    request.session['connect_tenant_color'] = tenant.primary_color
    request.session['connect_tenant_name'] = tenant.name
    request.session['tenant_id'] = tenant.id
    request.session['admin_from_panel'] = True  # Admin panelinden bağlandı
    
    # Superuser olarak giriş yap
    auth_logout(request)  # Önce mevcut kullanıcıdan çıkış yap
    login(request, superuser, backend='django.contrib.auth.backends.ModelBackend')
    
    messages.success(request, f'✅ "{tenant.name}" firmasına superuser olarak giriş yaptınız.')
    return redirect('home')

@login_required
@root_admin_required
def delete_tenant(request, tenant_id):
    """Firma silme - Tüm ilişkili verilerle birlikte silinir"""
    tenant = get_object_or_404(Tenant, id=tenant_id)
    tenant_name = tenant.name
    
    if request.method == 'POST':
        try:
            # Firma adını kaydet (mesaj için)
            tenant_name = tenant.name
            
            # Firmayı sil (CASCADE ilişkileri sayesinde tüm ilişkili veriler de silinir)
            tenant.delete()
            
            messages.success(request, f'✅ Firma "{tenant_name}" başarıyla silindi.')
        except Exception as e:
            messages.error(request, f'❌ Firma silinirken hata oluştu: {str(e)}')
        
        return redirect('admin_home')
    
    # GET request - Silme onay sayfası
    # Bu sayfayı kullanmak yerine JavaScript confirm kullanacağız
    return redirect('admin_home')

@login_required
@root_admin_required
def create_missing_admin_users(request):
    """Eksik admin kullanıcılarını oluştur"""
    User = get_user_model()
    root_admin = get_root_admin_user()
    
    if not root_admin:
        messages.error(request, 'Root admin kullanıcısı bulunamadı!')
        return redirect('admin_home')
    
    admin_password_hash = root_admin.password
    tenants = Tenant.objects.all()
    created_count = 0
    updated_count = 0
    
    for tenant in tenants:
        # Bu firma için admin kullanıcısı var mı kontrol et
        admin_user = User.objects.filter(
            tenant=tenant,
            user_code='admin',
            authority='Admin'
        ).first()
        
        if admin_user:
            # Admin kullanıcısı var, şifresini güncelle (root admin ile senkronize)
            if admin_user.password != admin_password_hash:
                admin_user.password = admin_password_hash
                admin_user.save(update_fields=['password'])
                updated_count += 1
        else:
            # Admin kullanıcısı yok, oluştur
            admin_username = f"{tenant.slug}_admin"
            
            # Eğer bu username zaten varsa, farklı bir username kullan
            counter = 1
            original_admin_username = admin_username
            while User.objects.filter(username=admin_username).exists():
                admin_username = f"{original_admin_username}_{counter}"
                counter += 1
            
            # Admin kullanıcısını oluştur
            admin_user = User.objects.create(
                username=admin_username,
                user_code='admin',
                first_name='Admin',
                last_name=tenant.name,
                email=tenant.email or f'admin@{tenant.slug}.fieldops.com',
                tenant=tenant,
                authority='Admin',
                is_staff=True,
                is_active=True
            )
            
            # Ana admin'in şifre hash'ini direkt atayalım
            admin_user.password = admin_password_hash
            admin_user.save(update_fields=['password'])
            
            # Admin rolü oluştur (eğer yoksa)
            admin_role, _ = UserRole.objects.get_or_create(
                name='Admin',
                tenant=tenant,
                defaults={'description': 'Firma yöneticisi'}
            )
            admin_user.role = admin_role
            admin_user.save(update_fields=['role'])
            
            created_count += 1
    
    if created_count > 0:
        messages.success(request, f'✅ {created_count} firma için admin kullanıcısı oluşturuldu.')
    if updated_count > 0:
        messages.info(request, f'🔄 {updated_count} firma için admin şifresi güncellendi.')
    if created_count == 0 and updated_count == 0:
        messages.info(request, 'ℹ️ Tüm firmalar için admin kullanıcısı zaten mevcut.')
    
    return redirect('admin_home')

# --- FİRMA SEÇME (DEVRE DIŞI) ---
# Bu özellik kaldırıldı. Admin panelinden firma seçimi yapılamaz.
# Kullanıcılar ana sayfadan firma adı girerek bağlanmalıdır.
@login_required
def select_company(request, tenant_id):
    """
    Bu özellik devre dışı bırakıldı.
    Admin panelinden firma seçimi artık yapılamaz.
    Kullanıcılar ana sayfadan (index.html) firma adı girerek bağlanmalıdır.
    """
    messages.info(request, 'Admin panelinden firma seçimi yapılamaz. Lütfen ana sayfadan firma adı girerek bağlanın.')
    return redirect('admin_home')

# --- ÖZEL GİRİŞ VIEW (Firma Adı ile) ---
from django.contrib.auth.views import LoginView
from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_protect
from django.utils.decorators import method_decorator
from django.urls import reverse_lazy

@method_decorator(csrf_protect, name='dispatch')
class CustomMobileLoginView(LoginView):
    """Mobil cihazlar için login view - CustomLoginView'a benzer ama mobile_home'a yönlendirir"""
    template_name = 'mobile/login.html'
    redirect_authenticated_user = False
    
    def get(self, request, *args, **kwargs):
        # Mobil cihaz kontrolü - eğer mobil değilse normal login'e yönlendir
        if not _is_mobile_device(request):
            # Tenant bilgisini session'dan al ve normal login'e yönlendir
            tenant_id = request.session.get('connect_tenant_id')
            tenant_slug = request.session.get('connect_tenant_slug')
            if tenant_id and tenant_slug:
                return redirect('login')
            return redirect('index')
        
        # Tenant bilgisini session'dan al
        tenant_id = request.session.get('connect_tenant_id')
        tenant_slug = request.session.get('connect_tenant_slug')
        
        # Eğer tenant bilgisi yoksa ana sayfaya yönlendir
        if not tenant_id or not tenant_slug:
            messages.error(request, 'Lütfen önce firma adını girin.')
            return redirect('mobile_index')
        
        try:
            tenant = Tenant.objects.get(id=tenant_id, slug=tenant_slug, is_active=True)
        except Tenant.DoesNotExist:
            for key in ['connect_tenant_id', 'connect_tenant_slug', 'connect_tenant_color', 'connect_tenant_name']:
                request.session.pop(key, None)
            messages.error(request, 'Firma bulunamadı. Lütfen tekrar deneyin.')
            return redirect('mobile_index')
        
        context = {
            'tenant': tenant,
            'primary_color': request.session.get('connect_tenant_color', tenant.primary_color),
        }
        return render(request, 'mobile/login.html', context)
    
    def post(self, request, *args, **kwargs):
        # Mobil cihaz kontrolü - eğer mobil değilse normal login'e yönlendir
        if not _is_mobile_device(request):
            # Tenant bilgisini session'dan al ve normal login'e yönlendir
            tenant_id = request.session.get('connect_tenant_id')
            tenant_slug = request.session.get('connect_tenant_slug')
            if tenant_id and tenant_slug:
                return redirect('login')
            return redirect('index')
        
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        # Tenant bilgisini session'dan al
        tenant_id = request.session.get('connect_tenant_id')
        tenant_slug = request.session.get('connect_tenant_slug')
        
        if not tenant_id or not tenant_slug:
            messages.error(request, 'Lütfen önce firma adını girin.')
            return redirect('mobile_index')
        
        try:
            tenant = Tenant.objects.get(id=tenant_id, slug=tenant_slug, is_active=True)
        except Tenant.DoesNotExist:
            messages.error(request, 'Firma bulunamadı. Lütfen tekrar deneyin.')
            return redirect('mobile_index')
        
        User = get_user_model()
        root_admin = get_root_admin_user()
        
        # Root admin kullanıcı adı ve şifresi ile giriş yapılıyorsa, o firmanın admin kullanıcısına giriş yap
        if root_admin and username == root_admin.username:
            if root_admin.check_password(password):
                # O firmanın admin kullanıcısını bul
                tenant_admin = User.objects.filter(
                    tenant=tenant,
                    user_code='admin',
                    authority='Admin'
                ).first()
                
                if not tenant_admin:
                    tenant_admin = User.objects.filter(
                        username__startswith=f"{tenant.slug}_admin",
                        tenant=tenant
                    ).first()
                
                if not tenant_admin:
                    tenant_admin = User.objects.filter(
                        tenant=tenant,
                        user_code='admin'
                    ).first()
                
                if tenant_admin:
                    login(request, tenant_admin, backend='django.contrib.auth.backends.ModelBackend')
                    request.session['tenant_id'] = tenant.id
                    request.session['connect_tenant_id'] = tenant.id
                    request.session['connect_tenant_slug'] = tenant.slug
                    request.session['connect_tenant_color'] = tenant.primary_color
                    request.session['connect_tenant_name'] = tenant.name
                    messages.success(request, f'"{tenant.name}" firmasına başarıyla giriş yaptınız.')
                    return redirect('mobile_home')
                else:
                    # Admin kullanıcısı yoksa otomatik oluştur
                    admin_username = f"{tenant.slug}_admin"
                    counter = 1
                    original_admin_username = admin_username
                    while User.objects.filter(username=admin_username).exists():
                        admin_username = f"{original_admin_username}_{counter}"
                        counter += 1
                    
                    tenant_admin = User.objects.create(
                        username=admin_username,
                        user_code='admin',
                        first_name='Admin',
                        last_name=tenant.name,
                        email=tenant.email or f'admin@{tenant.slug}.fieldops.com',
                        tenant=tenant,
                        authority='Admin',
                        is_staff=True,
                        is_active=True
                    )
                    tenant_admin.password = root_admin.password
                    tenant_admin.save(update_fields=['password'])
                    
                    admin_role, _ = UserRole.objects.get_or_create(
                        name='Admin',
                        tenant=tenant,
                        defaults={'description': 'Firma yöneticisi'}
                    )
                    tenant_admin.role = admin_role
                    tenant_admin.save(update_fields=['role'])
                    
                    login(request, tenant_admin, backend='django.contrib.auth.backends.ModelBackend')
                    request.session['tenant_id'] = tenant.id
                    request.session['connect_tenant_id'] = tenant.id
                    request.session['connect_tenant_slug'] = tenant.slug
                    request.session['connect_tenant_color'] = tenant.primary_color
                    request.session['connect_tenant_name'] = tenant.name
                    messages.success(request, f'"{tenant.name}" firması için admin kullanıcısı otomatik olarak oluşturuldu ve giriş yapıldı.')
                    return redirect('mobile_home')
            else:
                messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
                return redirect('mobile_login')
        
        # Normal kullanıcı girişi
        user = None
        username_attempts = [
            f"{tenant.slug}_{username}",
            username,
        ]
        
        for username_attempt in username_attempts:
            user = authenticate(request, username=username_attempt, password=password)
            if user is not None:
                break
        
        if user is not None and user.is_active:
            from apps.users.utils import is_root_admin
            
            if not is_root_admin(user):
                if hasattr(user, 'tenant') and user.tenant:
                    if user.tenant != tenant:
                        messages.error(request, f'Bu kullanıcı "{user.tenant.name}" firmasına aittir. Lütfen doğru firmayı seçin.')
                        return redirect('mobile_login')
                else:
                    user.tenant = tenant
                    user.save()
                    messages.info(request, f'Kullanıcınız "{tenant.name}" firmasına atandı.')
            
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            request.session['tenant_id'] = tenant.id
            request.session['connect_tenant_id'] = tenant.id
            request.session['connect_tenant_slug'] = tenant.slug
            request.session['connect_tenant_color'] = tenant.primary_color
            request.session['connect_tenant_name'] = tenant.name
            messages.success(request, f'"{tenant.name}" firmasına başarıyla giriş yaptınız.')
            return redirect('mobile_home')
        
        # Kullanıcı bulunamadı
        user_exists = False
        for username_attempt in username_attempts:
            try:
                potential_user = User.objects.get(username=username_attempt)
                if potential_user.tenant == tenant or (not hasattr(potential_user, 'tenant') or potential_user.tenant is None):
                    user_exists = True
                    break
            except User.DoesNotExist:
                continue
        
        if user_exists:
            messages.error(request, 'Şifre hatalı. Lütfen şifrenizi kontrol edin.')
        else:
            messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
        
        return redirect('mobile_login')

@method_decorator(csrf_protect, name='dispatch')
class CustomLoginView(LoginView):
    template_name = 'registration/login_tenant.html'
    redirect_authenticated_user = False  # Her zaman giriş yapılmasını iste
    
    def dispatch(self, request, *args, **kwargs):
        # Admin panelinden bağlanıldıysa ve session'da tenant varsa direkt geçir
        # Bu sadece admin panelinden yapılabilir
        admin_from_panel = request.session.get('admin_from_panel', False)
        connect_tenant_id = request.session.get('connect_tenant_id')
        
        if request.user.is_authenticated and admin_from_panel and connect_tenant_id:
            # Root admin kontrolü - sadece admin panelinden bağlanıldıysa geçir
            if is_root_admin(request.user):
                request.session['tenant_id'] = connect_tenant_id
                from django.urls import reverse
                return redirect(reverse('home'))
        
        # Normal kullanıcılar için her zaman login yapılması gerekiyor
        # Admin panelinden değilse, authenticated olsa bile login sayfasını göster
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        # redirect_authenticated_user = False olduğu için authenticated kullanıcılar da login sayfasını görebilir
        # Bu sayede her zaman giriş yapılması zorunlu olur
        
        # Tenant bilgisini session'dan al
        tenant_id = request.session.get('connect_tenant_id')
        tenant_slug = request.session.get('connect_tenant_slug') or self.kwargs.get('tenant_slug')

        # KURAL: Firma adıyla bağlanılmadan login ekranından sistem seçimi YOK.
        if not tenant_id or not tenant_slug:
            messages.error(request, 'Lütfen önce firma adını girip "Bağlan" butonuna basın.')
            return redirect('index')

        try:
            tenant = Tenant.objects.get(id=tenant_id, slug=tenant_slug, is_active=True)
        except Tenant.DoesNotExist:
            for key in ['connect_tenant_id', 'connect_tenant_slug', 'connect_tenant_color', 'connect_tenant_name', 'admin_from_panel']:
                request.session.pop(key, None)
            return redirect('index')

        # Logout sonrası gelindi mi kontrol et
        from_tenant_logout = request.session.get('from_tenant_logout', False)
        logout_tenant_name = request.session.get('logout_tenant_name', tenant.name)

        context = {
            'tenant': tenant,
            'primary_color': request.session.get('connect_tenant_color', tenant.primary_color),
            'from_tenant_logout': from_tenant_logout,
            'logout_tenant_name': logout_tenant_name,
        }
        return render(request, 'registration/login_tenant.html', context)
    
    def post(self, request, *args, **kwargs):
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        User = get_user_model()
        root_admin = get_root_admin_user()

        # Tenant'ı sadece session'dan al (firma adıyla bağlanılmadan login yok)
        tenant = None
        tenant_id = request.session.get('connect_tenant_id')
        tenant_slug = request.session.get('connect_tenant_slug')
        if tenant_id:
            try:
                tenant = Tenant.objects.get(id=tenant_id, is_active=True)
                if tenant_slug and tenant.slug != tenant_slug:
                    messages.error(request, 'Firma bilgisi uyumsuz. Lütfen tekrar firma adını girin.')
                    return redirect('index')
            except Tenant.DoesNotExist:
                messages.error(request, 'Firma bulunamadı. Lütfen tekrar firma adını girin.')
                for key in ['connect_tenant_id', 'connect_tenant_slug', 'connect_tenant_color', 'connect_tenant_name']:
                    request.session.pop(key, None)
                return redirect('index')
        else:
            messages.error(request, 'Lütfen önce firma adını girip "Bağlan" butonuna basın.')
            return redirect('index')
        
        # Tenant bulunamadıysa hata ver
        if not tenant:
            messages.error(request, 'Firma bilgisi bulunamadı. Lütfen tekrar deneyin.')
            return redirect('index')
        
        # Root admin değilse, normal kullanıcı girişi yap
        # Firma seçildi, normal kullanıcı girişi yapılabilir
        user = None
        username_attempts = [
            f"{tenant.slug}_{username}",  # tenant_slug_username
            username,  # username (eski kullanıcılar için)
        ]
        
        for username_attempt in username_attempts:
            user = authenticate(request, username=username_attempt, password=password)
            if user is not None:
                break
        
        if user is not None and user.is_active:
            from apps.users.utils import is_root_admin

            # KURAL 1: Root admin kullanıcı sadece Rotexia tenant'ında giriş yapabilir.
            if is_root_admin(user):
                # Rotexia tenant'ı kontrolü (slug veya name ile)
                rotexia_slugs = ['rotexia', 'sistem', 'admin']
                if tenant.slug.lower() not in rotexia_slugs and tenant.name.lower() not in ['rotexia', 'sistem', 'admin']:
                    messages.error(
                        request,
                        'Bu kullanıcı sadece "Rotexia" sistemi için kullanılabilir. '
                        'Lütfen firma adını "Rotexia" yazıp bağlanarak giriş yapın.'
                    )
                    return redirect('index')

            # KURAL 2: Tüm tenant kullanıcıları (superuser dahil) kesinlikle kendi tenant'ı dışında giriş yapamaz.
            # ÖNEMLİ: Tenant'a bağlı superuser'lar da (pasteladmin gibi) root admin DEĞİLDİR!
            if not is_root_admin(user):
                # Kullanıcının tenant'ı yoksa hata ver
                if not hasattr(user, 'tenant') or not user.tenant:
                    messages.error(request, 'Bu kullanıcı bir firmaya atanmamış. Lütfen yöneticiniz ile iletişime geçin.')
                    return redirect('index')
                
                # Kullanıcının tenant'ı seçilen tenant ile eşleşmeli
                if user.tenant != tenant:
                    messages.error(request, f'Bu kullanıcı "{user.tenant.name}" firmasına aittir. Lütfen doğru firmayı seçin.')
                    return redirect('index')
            
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            # Session'a tenant bilgisini kaydet
            request.session['tenant_id'] = tenant.id
            request.session['connect_tenant_id'] = tenant.id
            request.session['connect_tenant_slug'] = tenant.slug
            request.session['connect_tenant_color'] = tenant.primary_color
            request.session['connect_tenant_name'] = tenant.name
            messages.success(request, f'"{tenant.name}" firmasına başarıyla giriş yaptınız.')
            return redirect('home')
        
        # Kullanıcı bulunamadı veya şifre yanlış
        # Kullanıcı adı var mı kontrol et
        user_exists = False
        for username_attempt in username_attempts:
            try:
                potential_user = User.objects.get(username=username_attempt)
                if potential_user.tenant == tenant or (not hasattr(potential_user, 'tenant') or potential_user.tenant is None):
                    user_exists = True
                    break
            except User.DoesNotExist:
                continue
        
        if user_exists:
            messages.error(request, 'Şifre hatalı. Lütfen şifrenizi kontrol edin.')
        else:
            messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
        
        return redirect('login')

# --- GELİŞTİRİCİ ADMIN GİRİŞİ ---
@method_decorator(csrf_protect, name='dispatch')
class AdminLoginView(LoginView):
    template_name = 'registration/admin_login.html'
    redirect_authenticated_user = False  # Her zaman giriş yapılmasını iste

    def post(self, request, *args, **kwargs):
        """Root admin girişi: sadece root admin kullanıcıları kabul edilir."""
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user is None or not user.is_active:
            messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
            return self.form_invalid(self.get_form())

        if not is_root_admin(user):
            messages.error(request, 'Bu giriş ekranı sadece ana sistem (Rotexia) yöneticisi içindir.')
            return self.form_invalid(self.get_form())

        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        # Admin girişinde tenant session'ını temizle (kesin kural)
        for key in ['tenant_id', 'connect_tenant_id', 'connect_tenant_slug', 'connect_tenant_color', 'connect_tenant_name', 'admin_from_panel']:
            request.session.pop(key, None)
        return redirect(self.get_success_url())
    
    def get_success_url(self):
        # Root admin kontrolü
        if is_root_admin(self.request.user):
            return reverse_lazy('admin_home')
        return reverse_lazy('home')

# --- MOBİL ANASAYFA ---
@login_required
def mobile_home(request):
    today = date.today()
    user = request.user
    # Supervisor (veya hiyerarşide altı olan herkes) için sekme göster
    scope_no_self = get_hierarchy_scope_for_user(user, include_self=False)
    # Not: HierarchyScope.usernames boş ise (Admin/superuser) "herkesi görür" anlamına gelir.
    has_team = bool(scope_no_self.usernames) or getattr(user, "is_superuser", False) or getattr(user, "authority", None) == "Admin"
    
    # Tenant kontrolü - sadece kullanıcının tenant'ına ait verileri göster
    tenant = getattr(request, 'tenant', None)
    if not tenant and hasattr(user, 'tenant') and user.tenant:
        tenant = user.tenant
    
    # Aktif ziyaret var mı kontrol et (check_in_time var ama check_out_time yok) - tenant filtresi ile
    active_visit_qs = VisitTask.objects.filter(
        merch_code=user.username,
        check_in_time__isnull=False,
        check_out_time__isnull=True,
        status__in=['pending', 'missed']  # completed değilse aktif
    )
    if tenant:
        active_visit_qs = active_visit_qs.filter(tenant=tenant)
    active_visit = active_visit_qs.first()
    
    # Eğer aktif ziyaret varsa uyarı göster
    if active_visit:
        from django.contrib import messages
        messages.warning(request, f'Devam eden bir ziyaret var: {active_visit.customer.name}. Lütfen önce bu ziyareti tamamlayın.')
    
    # Görevleri al - tenant filtresi ile
    tasks_qs = VisitTask.objects.filter(
        merch_code=user.username,
        planned_date=today
    )
    if tenant:
        tasks_qs = tasks_qs.filter(tenant=tenant)
    tasks = tasks_qs.select_related('customer').order_by('status', 'customer__name')
    
    total_tasks = tasks.count()
    completed_tasks = tasks.filter(status='completed').count()
    
    progress_percentage = 0
    if total_tasks > 0:
        progress_percentage = int((completed_tasks / total_tasks) * 100)
    
    context = {
        'tasks': tasks,
        'today': today,
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'progress_percentage': progress_percentage,
        'active_visit': active_visit,
        'has_team': has_team,
    }
    return render(request, 'mobile/home.html', context)


@login_required
def mobile_team_home(request):
    """
    Supervisor tab: shows team progress (today) and subordinate list.
    """
    today = date.today()
    user = request.user
    ensure_root_admin_configured(user)

    # Root admin: "supervisor gibi görüntüle" desteği
    as_user_id = request.GET.get("as_user")
    if is_root_admin(user) and as_user_id:
        try:
            subject = get_user_model().objects.get(id=int(as_user_id))
        except Exception:
            subject = user
    else:
        subject = user

    # Root admin ana görünüm: Admin node altındaki tüm atanmış kullanıcılar
    if is_root_admin(user) and subject == user and not as_user_id:
        user_ids = get_assigned_user_ids_under_admin_node()
        if user_ids:
            sub_users = list(
                get_user_model().objects.filter(id__in=user_ids).order_by("first_name", "last_name", "username")
            )
            sub_codes = [u.username for u in sub_users]
        else:
            # Fallback: Admin her şeyi görebilir (hiyerarşi ataması yoksa bile)
            sub_users = list(
                get_user_model().objects.exclude(id=user.id).order_by("first_name", "last_name", "username")
            )
            sub_codes = [u.username for u in sub_users]
    else:
        scope_no_self = get_hierarchy_scope_for_user(subject, include_self=False)
        # Admin/superuser için boş set => herkesi görür
        if not scope_no_self.usernames and (
            getattr(subject, "is_superuser", False) or getattr(subject, "authority", None) == "Admin"
        ):
            sub_users = list(
                get_user_model().objects.exclude(id=subject.id).order_by("first_name", "last_name", "username")
            )
            sub_codes = [u.username for u in sub_users]
        else:
            sub_codes = sorted(scope_no_self.usernames)
            sub_users = list(
                get_user_model().objects.filter(username__in=sub_codes).order_by("first_name", "last_name", "username")
            )

    stats_qs = (
        VisitTask.objects.filter(planned_date=today, merch_code__in=sub_codes)
        .values("merch_code")
        .annotate(
            total=Count("id"),
            completed=Count("id", filter=DQ(status="completed")),
        )
    )
    stats = {x["merch_code"]: x for x in stats_qs}

    team_total = sum(x["total"] for x in stats.values()) if stats else 0
    team_completed = sum(x["completed"] for x in stats.values()) if stats else 0
    team_progress_percentage = int((team_completed / team_total) * 100) if team_total > 0 else 0

    team = []
    for u in sub_users:
        s = stats.get(u.username, {"total": 0, "completed": 0})
        # Konum için: aktif ziyaret varsa onun mağazası, yoksa bugünkü ilk görev
        active = (
            VisitTask.objects.filter(
                merch_code=u.username,
                check_in_time__isnull=False,
                check_out_time__isnull=True,
                status__in=["pending", "missed"],
            )
            .select_related("customer")
            .order_by("-check_in_time")
            .first()
        )
        candidate = active
        if not candidate:
            candidate = (
                VisitTask.objects.filter(merch_code=u.username, planned_date=today)
                .select_related("customer")
                .order_by("status", "customer__name")
                .first()
            )
        location_url = None
        if candidate and candidate.customer and candidate.customer.latitude and candidate.customer.longitude:
            location_url = f"https://www.google.com/maps/search/?api=1&query={candidate.customer.latitude},{candidate.customer.longitude}"

        name = (f"{u.first_name} {u.last_name}").strip() or u.username
        # Root admin için: bu kişinin ekip ekranına drill-down mümkün mü?
        drill_as_user_id = None
        if is_root_admin(user):
            sc = get_hierarchy_scope_for_user(u, include_self=False)
            if sc.usernames:
                drill_as_user_id = u.id
        team.append(
            {
                "user_id": u.id,
                "name": name,
                "user_code": getattr(u, "user_code", u.username),
                "merch_code": u.username,
                "authority": getattr(u, "authority", ""),
                "completed": s["completed"],
                "total": s["total"],
                "location_url": location_url,
                "drill_as_user_id": drill_as_user_id,
            }
        )

    return render(
        request,
        "mobile/team_home.html",
        {
            "today": today,
            "has_team": True,
            "team": team,
            "team_total_tasks": team_total,
            "team_completed_tasks": team_completed,
            "team_progress_percentage": team_progress_percentage,
            "viewing_as": (None if subject == user else subject),
            "is_root_admin_view": is_root_admin(user),
        },
    )

# --- MOBİL PROFİL (Hata veren eksik parça buydu) ---
@login_required
def mobile_profile(request):
    return render(request, 'mobile/profile.html')

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from apps.core.excel_utils import xlsx_from_rows

@login_required
def download_excel_template(request, template_type):
    """
    İstenilen türe göre (customer, user, task) boş bir Excel şablonu oluşturur ve indirir.
    HATA DÜZELTME: Tüm sütunların uzunluğu eşitlendi.
    """
    filename = "sablon.xlsx"
    columns = []
    
    # 1. Sütun Başlıklarını Belirle
    if template_type == 'customer':
        columns = [
            'Müşteri Kodu', 'Müşteri Adı', 'Cari / Firma', 'İl', 'İlçe', 
            'Adres', 'Telefon', 'Yetkili Kişi', 'Enlem', 'Boylam'
        ]
        filename = "musteri_yukleme_sablonu.xlsx"

    elif template_type == 'user':
        columns = ['Kullanıcı Kodu', 'Ad', 'Soyad', 'Telefon', 'E-posta', 'Rol', 'Şifre']
        filename = "personel_yukleme_sablonu.xlsx"

    elif template_type == 'task':
        columns = ['Müşteri Kodu', 'Personel', 'Tarih', 'Ziyaret Notu']
        filename = "gorev_yukleme_sablonu.xlsx"
        
    elif template_type == 'route':
        columns = ['Saha Kullanıcısı', 'Müşteri Kodu', 'Gün 1', 'Gün 2', 'Gün 3', 'Gün 4', 'Gün 5', 'Gün 6', 'Gün 7']
        filename = "rota_yukleme_sablonu.xlsx"

    # Tek satırlık örnek şablon
    row = {col: "" for col in columns}

    if template_type == 'customer':
        row['Müşteri Kodu'] = 'M-001'
        row['Müşteri Adı'] = 'Örnek Market'
        row['İl'] = 'İstanbul'
    elif template_type == 'user':
        row['Kullanıcı Kodu'] = 'Merch1'
        row['Ad'] = 'Ahmet'
        row['Rol'] = 'Saha Personeli'
        row['Şifre'] = '123456'
    elif template_type == 'task':
        row['Müşteri Kodu'] = 'M-001'
        row['Personel'] = 'Merch1'
        row['Tarih'] = '25.12.2025'
    elif template_type == 'route':
        row['Saha Kullanıcısı'] = 'Merch1'
        row['Müşteri Kodu'] = 'M-001'
        row['Gün 1'] = '1'

    content = xlsx_from_rows([row], sheet_name="Şablon", header_order=columns)
    response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

from django.shortcuts import get_object_or_404
from apps.forms.models import Survey  # Bunu en üste eklemeyi unutma

@login_required
def mobile_task_detail(request, pk):
    """
    Seçilen görevin detay ekranı.
    Mağaza bilgisi + Formlar + Başlat Butonu
    Filtreleme: Tüm filtreler AND mantığıyla çalışır (şartlı)
    """
    user = request.user
    
    # Tenant kontrolü - sadece kullanıcının tenant'ına ait görevleri göster
    tenant = getattr(request, 'tenant', None)
    if not tenant and hasattr(user, 'tenant') and user.tenant:
        tenant = user.tenant
    
    # Tenant filtresi ile task'ı al
    if tenant:
        task = get_object_or_404(VisitTask, pk=pk, tenant=tenant)
    else:
        task = get_object_or_404(VisitTask, pk=pk)
        # Tenant yoksa ama task'ın tenant'ı varsa erişim yok
        if hasattr(task, 'tenant') and task.tenant:
            return HttpResponseForbidden("Bu görevi görüntüleme yetkiniz yok.")
    customer = task.customer
    # Görevin sahibi (form filtreleri ve cevap görüntüleme için)
    task_user = get_user_model().objects.filter(username=task.merch_code).first() or user

    # Yetki kontrolü: görevi sadece sahibi veya hiyerarşide üstü görebilir.
    scope = get_hierarchy_scope_for_user(user, include_self=True)
    if scope.usernames and task.merch_code not in scope.usernames:
        return HttpResponseForbidden("Bu görevi görüntüleme yetkiniz yok.")
    # Aksiyon (ziyaret başlat/bitir, form doldur) sadece görevin sahibinde olmalı
    can_act = (task.merch_code == user.username)
    
    # Tenant kontrolü - sadece kullanıcının tenant'ına ait anketleri göster
    tenant = getattr(request, 'tenant', None)
    if not tenant and hasattr(user, 'tenant') and user.tenant:
        tenant = user.tenant
    
    # Aktif anketleri başlangıç olarak al - sadece bu tenant'a ait (temel yapı yok, anketler firma-özel)
    if tenant:
        surveys = Survey.objects.filter(is_active=True, tenant=tenant)
    else:
        # Tenant yoksa boş liste (güvenlik)
        surveys = Survey.objects.none()
    
    # Tarih kontrolü
    from datetime import date
    today = date.today()
    surveys = surveys.filter(
        models.Q(start_date__isnull=True) | models.Q(start_date__lte=today)
    ).filter(
        models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
    )
    
    # FİLTRELEME (AND MANTIĞI - TÜM ŞARTLAR SAĞLANMALI)
    filtered_surveys = []
    
    for survey in surveys:
        should_show = True  # Varsayılan olarak göster
        
        # 1. KULLANICI FİLTRESİ
        if survey.filter_users.exists():
            # Eğer kullanıcı filtresi varsa, kullanıcı listede olmalı
            if task_user not in survey.filter_users.all():
                should_show = False
        
        # 2. ROL FİLTRESİ
        if survey.target_roles.exists():
            # Eğer rol filtresi varsa, kullanıcının rolü listede olmalı
            if not task_user.role or task_user.role not in survey.target_roles.all():
                should_show = False
        
        # 3. KULLANICI ÖZEL ALAN FİLTRELERİ
        if survey.user_custom_filters:
            for field_slug, allowed_values in survey.user_custom_filters.items():
                if allowed_values:  # Eğer değer seçilmişse
                    user_value_str = task_user.extra_data.get(field_slug, '') if task_user.extra_data else ''
                    # Tag sistemi: Değerler virgülle ayrılmış (örn: "Lansman,Stok Takibi")
                    user_tags = [tag.strip() for tag in str(user_value_str).split(',') if tag.strip()]
                    # Kullanıcının tag'lerinden en az biri, izin verilen değerlerden biri olmalı
                    if not any(tag in allowed_values for tag in user_tags):
                        should_show = False
                        break
        
        # 4. MÜŞTERİ FİLTRESİ
        if survey.filter_customers.exists():
            # Eğer müşteri filtresi varsa, müşteri listede olmalı
            if customer not in survey.filter_customers.all():
                should_show = False
        
        # 5. CARİ FİLTRESİ
        if survey.filter_caris.exists():
            # Eğer cari filtresi varsa, müşterinin carisi listede olmalı
            if not customer.cari or customer.cari not in survey.filter_caris.all():
                should_show = False
        
        # 6. MÜŞTERİ ÖZEL ALAN FİLTRELERİ
        if survey.custom_filters:
            for field_slug, allowed_values in survey.custom_filters.items():
                if allowed_values:  # Eğer değer seçilmişse
                    customer_value = customer.extra_data.get(field_slug, '') if customer.extra_data else ''
                    # Müşterinin bu alandaki değeri, izin verilen değerlerden biri olmalı
                    if customer_value not in allowed_values:
                        should_show = False
                        break
        
        # Tüm şartlar sağlandıysa listeye ekle
        if should_show:
            filtered_surveys.append(survey)
    
    # --- Anket durumları (Yapıldı / Yapılmadı) ---
    # Kural:
    # - Anketin zorunlu soruları varsa: tüm zorunlular cevaplandıysa "Yapıldı"
    # - Zorunlu soru yoksa: ankete ait en az 1 soru cevaplandıysa "Yapıldı"
    survey_ids = [s.id for s in filtered_surveys]
    answered_q_ids = set(
        SurveyAnswer.objects.filter(task=task, question__survey_id__in=survey_ids)
        .exclude(answer_text__isnull=True, answer_photo__isnull=True)
        .values_list('question_id', flat=True)
    )

    questions = list(
        Question.objects.filter(survey_id__in=survey_ids).values('id', 'survey_id', 'required')
    )
    req_map = {}
    all_map = {}
    for q in questions:
        sid = q['survey_id']
        all_map.setdefault(sid, []).append(q['id'])
        if q['required']:
            req_map.setdefault(sid, []).append(q['id'])

    for s in filtered_surveys:
        req_ids = req_map.get(s.id, [])
        all_ids = all_map.get(s.id, [])
        if req_ids:
            done_required = sum(1 for qid in req_ids if qid in answered_q_ids)
            s.is_done = done_required == len(req_ids)
            s.required_done = done_required
            s.required_total = len(req_ids)
        else:
            any_done = any(qid in answered_q_ids for qid in all_ids)
            s.is_done = any_done
            s.required_done = 1 if any_done else 0
            s.required_total = 1 if all_ids else 0
        # Üst kullanıcılar için read-only görüntüleme: en az 1 cevap var mı?
        s.has_answers = any(qid in answered_q_ids for qid in all_ids)

    context = {
        'task': task,
        'surveys': filtered_surveys,
        'can_act': can_act,
    }
    return render(request, 'mobile/task_detail.html', context)


@login_required
def mobile_view_survey(request, task_id, survey_id):
    """
    Read-only: Üst kullanıcılar (hiyerarşide) ve görev sahibi, doldurulan anketi görüntüleyebilir.
    Düzenleme yoktur.
    """
    task = get_object_or_404(VisitTask, pk=task_id)
    user = request.user
    
    # Tenant kontrolü - sadece kullanıcının tenant'ına ait anketleri göster
    tenant = getattr(request, 'tenant', None)
    if not tenant and hasattr(user, 'tenant') and user.tenant:
        tenant = user.tenant
    
    # Tenant filtresi ile survey'yi al
    if tenant:
        survey = get_object_or_404(Survey, pk=survey_id, tenant=tenant)
    else:
        # Tenant yoksa erişim yok
        return HttpResponseForbidden("Bu formu görüntüleme yetkiniz yok.")

    scope = get_hierarchy_scope_for_user(user, include_self=True)
    if scope.usernames and task.merch_code not in scope.usernames:
        return HttpResponseForbidden("Bu görevin formunu görüntüleme yetkiniz yok.")

    questions = list(survey.questions.all().order_by("order"))
    answers = list(
        SurveyAnswer.objects.filter(task=task, question__in=questions).select_related("question")
    )
    ans_by_qid = {a.question_id: a for a in answers}

    items = []
    for q in questions:
        a = ans_by_qid.get(q.id)
        items.append(
            {
                "question": q,
                "answer_text": (a.answer_text if a else None),
                "answer_photo": (a.answer_photo if a else None),
                "has_answer": bool(a and ((a.answer_text and str(a.answer_text).strip()) or a.answer_photo)),
            }
        )

    return render(
        request,
        "mobile/survey_view.html",
        {
            "task": task,
            "survey": survey,
            "items": items,
        },
    )

@login_required
def mobile_fill_survey(request, task_id, survey_id):
    task = get_object_or_404(VisitTask, pk=task_id)
    user = request.user
    
    # Tenant kontrolü - sadece kullanıcının tenant'ına ait anketleri göster
    tenant = getattr(request, 'tenant', None)
    if not tenant and hasattr(user, 'tenant') and user.tenant:
        tenant = user.tenant
    
    # Tenant filtresi ile survey'yi al
    if tenant:
        survey = get_object_or_404(Survey, pk=survey_id, tenant=tenant)
    else:
        # Tenant yoksa erişim yok
        return HttpResponseForbidden("Bu formu doldurma yetkiniz yok.")
    
    # Sadece görevin sahibi form doldurabilir
    if task.merch_code != user.username:
        return HttpResponseForbidden("Bu görevin formunu doldurma yetkiniz yok.")
    
    # Ana soruları al (parent_question veya dependency_question olmayanlar)
    main_questions = survey.questions.filter(
        models.Q(parent_question__isnull=True) & models.Q(dependency_question__isnull=True)
    ).order_by('order')
    
    # Tüm soruları al (alt sorular dahil)
    all_questions = survey.questions.all().order_by('order')

    if request.method == 'POST':
        # AJAX isteği mi kontrol et
        immediate = request.POST.get('immediate', 'false') == 'true'
        
        # Eğer immediate değilse, localStorage'a kaydet (JavaScript tarafında yapılacak)
        # Burada sadece normal form gönderimini işle
        try:
            # Her soru için döngüye girip cevabı alalım
            for q in all_questions:
                # HTML formundaki input ismi: "q_1", "q_2" şeklinde ayarlamıştık
                input_name = f"q_{q.id}"
                
                text_val = request.POST.get(input_name)
                photo_val = request.FILES.get(input_name)
                photo_b64 = request.POST.get(f"{input_name}_base64") if q.input_type == 'photo' else None

                # WebView fallback: base64 geldiyse dosyaya çevir
                if not photo_val and photo_b64 and photo_b64.startswith('data:image/'):
                    try:
                        header, b64data = photo_b64.split(',', 1)
                        # data:image/jpeg;base64
                        ext = 'jpg'
                        if 'image/png' in header:
                            ext = 'png'
                        elif 'image/webp' in header:
                            ext = 'webp'
                        elif 'image/jpeg' in header or 'image/jpg' in header:
                            ext = 'jpg'

                        import base64
                        decoded = base64.b64decode(b64data)
                        photo_val = ContentFile(decoded, name=f"survey_{task_id}_{q.id}.{ext}")
                    except Exception:
                        photo_val = None
                
                # Video kontrolü
                video_val = None
                if q.input_type == 'video':
                    video_file = request.FILES.get(f'q_{q.id}')
                    if video_file:
                        video_val = video_file
                
                # Konum kontrolü
                latitude = None
                longitude = None
                if q.input_type == 'location':
                    lat_str = request.POST.get(f'q_{q.id}_latitude', '').strip()
                    lng_str = request.POST.get(f'q_{q.id}_longitude', '').strip()
                    if lat_str and lng_str:
                        try:
                            latitude = float(lat_str)
                            longitude = float(lng_str)
                        except (ValueError, TypeError):
                            latitude = None
                            longitude = None
                
                # Eğer soruya bir cevap verilmişse (Yazı, Fotoğraf, Video veya Konum)
                if text_val or photo_val or video_val or (latitude is not None and longitude is not None):
                    # Önce eski cevap varsa silelim (Güncelleme mantığı)
                    SurveyAnswer.objects.filter(task=task, question=q).delete()
                    
                    # Yeni cevabı kaydet
                    SurveyAnswer.objects.create(
                        task=task,
                        question=q,
                        answer_text=text_val,
                        answer_photo=photo_val,
                        answer_video=video_val,
                        latitude=latitude,
                        longitude=longitude,
                        tenant=tenant
                    )
            
            # AJAX isteği ise JSON döndür
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax') == 'true':
                return JsonResponse({'success': True, 'message': 'Form başarıyla kaydedildi.'})
            
            messages.success(request, '✅ Form başarıyla kaydedildi.')
            return redirect('mobile_task_detail', pk=task_id)
            
        except Exception as e:
            # AJAX isteği ise JSON döndür
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax') == 'true':
                return JsonResponse({'success': False, 'message': f'Hata oluştu: {str(e)}'}, status=400)
            messages.error(request, f'Hata oluştu: {str(e)}')

    # Soruları ve alt sorularını hazırla
    questions_data = []
    for q in main_questions:
        # Alt soruları al (parent_question veya dependency_question ile bağlı olanlar)
        child_questions = survey.questions.filter(
            models.Q(parent_question=q) | models.Q(dependency_question=q)
        ).distinct().order_by('order')
        
        # Alt soruları detaylı bilgilerle hazırla
        child_questions_list = []
        for child in child_questions:
            # Hangi alan kullanılmış? (parent_question veya dependency_question)
            trigger_value = ''
            if child.parent_question == q:
                trigger_value = child.trigger_answer or ''
            elif child.dependency_question == q:
                trigger_value = child.dependency_value or ''
            
            child_questions_list.append({
                'id': child.id,
                'label': child.label,
                'input_type': child.input_type,
                'required': child.required,
                'trigger_answer': trigger_value,
                'parent_id': q.id,
                'options': [{'text': opt.text, 'id': opt.id} for opt in child.options.all()],
                'min_photos': child.min_photos,
                'max_photos': child.max_photos,
            })
        
        # Sorunun seçeneklerini al (select tipi sorular için)
        question_options = [{'text': opt.text, 'id': opt.id} for opt in q.options.all()] if q.input_type == 'select' else []
        
        questions_data.append({
            'question': q,
            'child_questions': child_questions_list,
            'options': question_options,
        })
    
    # Zorunlu soruları kontrol et
    required_questions = [q for q in all_questions if q.required]
    answered_questions = SurveyAnswer.objects.filter(task=task, question__in=required_questions).values_list('question_id', flat=True)
    missing_required = [q for q in required_questions if q.id not in answered_questions]

    context = {
        'task': task,
        'task_id': task.id,  # Template için task_id ekle
        'survey': survey,
        'survey_id': survey.id,  # Template için survey_id ekle
        'questions_data': questions_data,
        'required_questions': required_questions,
        'missing_required': missing_required,
    }
    return render(request, 'mobile/survey_form.html', context)

@csrf_exempt
@login_required
def start_visit_check(request, task_id):
    # Debug için print ekleyelim (production'da kaldırılabilir)
    print(f"[DEBUG] start_visit_check çağrıldı - Method: {request.method}, Task ID: {task_id}")
    print(f"[DEBUG] User: {request.user if hasattr(request, 'user') else 'Anonymous'}")
    print(f"[DEBUG] Body: {request.body}")
    print(f"[DEBUG] Headers: {dict(request.headers)}")
    
    # GET isteği için test endpoint
    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'message': 'API çalışıyor',
            'task_id': task_id,
            'method': 'GET'
        })
    
    if request.method == 'POST':
        try:
            # 1. Mobilden gelen body
            body_text = request.body.decode('utf-8')
            data = json.loads(body_text)

            # 2. Görevi ve Müşteriyi Bul
            task = VisitTask.objects.get(id=task_id)
            customer = task.customer
            # Sadece görevin sahibi ziyareti başlatabilir (Admin dahil istisna yok)
            if task.merch_code != getattr(request.user, "username", None):
                return JsonResponse({'success': False, 'message': 'Bu ziyareti başlatma yetkiniz yok.'}, status=403)

            # 3. Mesafe kuralı kontrolü (distance_rule) - kapalıysa konum ve mesafe kontrolü yapılmaz
            distance_rule_setting = SystemSetting.objects.filter(key='distance_rule').first()
            # Eski require_gps ayarını migrate et
            if not distance_rule_setting:
                old_setting = SystemSetting.objects.filter(key='require_gps').first()
                if old_setting:
                    old_setting.key = 'distance_rule'
                    old_setting.label = 'Mesafe Kuralı'
                    old_setting.description = 'Açık: Giriş mesafesi ve gezinme mesafesi kontrolü yapılır. Kapalı: Mesafe kontrolü yapılmaz, herhangi bir mesafeden ziyaret başlatılabilir.'
                    old_setting.save()
                    distance_rule_setting = old_setting

            distance_rule_enabled = True
            if distance_rule_setting:
                try:
                    distance_rule_enabled = distance_rule_setting.value.lower() == 'true'
                except:
                    distance_rule_enabled = True

            # Mesafe kuralı kapalıysa: konum zorunlu değil, doğrudan başlat
            if not distance_rule_enabled:
                from datetime import datetime
                task.check_in_time = datetime.now()
                task.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Ziyaret başlatıldı. (Mesafe kuralı kapalı olduğu için konum ve mesafe kontrolü yapılmadı.)'
                })

            # 4. Kullanıcı koordinatları - mesafe kuralı açıkken zorunlu
            user_lat_raw = data.get('latitude')
            user_lon_raw = data.get('longitude')

            if user_lat_raw is None or user_lon_raw is None:
                return JsonResponse({
                    'success': False, 
                    'message': 'Konum bilgisi alınamadı. Lütfen GPS\'in açık olduğundan ve konum izninin verildiğinden emin olun.'
                })

            try:
                user_lat = float(user_lat_raw)
                user_lon = float(user_lon_raw)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False, 
                    'message': 'Geçersiz konum bilgisi. Lütfen tekrar deneyin.'
                })

            # Koordinatlar geçerli aralıkta mı?
            if not (-90 <= user_lat <= 90) or not (-180 <= user_lon <= 180):
                return JsonResponse({
                    'success': False, 
                    'message': 'Geçersiz konum koordinatları. Lütfen tekrar deneyin.'
                })

            try:
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Koordinat alındı - Lat: {user_lat}, Lon: {user_lon}")
            except:
                pass

            try:
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Task ve Customer bulundu - Customer: {customer.name}, Lat: {customer.latitude}, Lon: {customer.longitude}")
            except:
                pass
            
            # Müşterinin koordinatı yoksa mesafe kontrolü yapmadan ziyareti başlat
            # latitude ve longitude FloatField, None veya 0.0 olabilir
            cust_lat_val = customer.latitude
            cust_lng_val = customer.longitude
            
            # Koordinat kontrolü - None veya 0.0 ise mesafe kontrolü yapma
            if (cust_lat_val is None or cust_lng_val is None or 
                cust_lat_val == 0.0 or cust_lng_val == 0.0):
                from datetime import datetime
                task.check_in_time = datetime.now()
                task.save()
                return JsonResponse({
                    'success': True, 
                    'message': 'Ziyaret başlatıldı. (Müşteri konumu sistemde olmadığı için mesafe kontrolü yapılmadı.)'
                })

            # 3. Mesafeyi Hesapla - None kontrolü ile
            try:
                # None kontrolü
                if cust_lat_val is None or cust_lng_val is None:
                    raise ValueError("Müşteri koordinatları None")
                
                cust_lat = float(cust_lat_val)
                cust_lon = float(cust_lng_val)
                
                # Koordinatlar geçerli mi kontrol et (enlem: -90 ile 90, boylam: -180 ile 180)
                if not (-90 <= cust_lat <= 90) or not (-180 <= cust_lon <= 180):
                    raise ValueError("Koordinatlar geçersiz aralıkta")
                
                distance = calculate_distance(user_lat, user_lon, cust_lat, cust_lon)
            except (ValueError, TypeError) as e:
                # Koordinat geçersizse mesafe kontrolü yapmadan başlat
                from datetime import datetime
                task.check_in_time = datetime.now()
                task.save()
                return JsonResponse({
                    'success': True, 
                    'message': 'Ziyaret başlatıldı. (Müşteri konumu geçersiz olduğu için mesafe kontrolü yapılmadı.)'
                })
            
            # 4. Mesafe Kuralı kontrolü - Eğer kapalıysa mesafe kontrolü yapma
            distance_rule_setting = SystemSetting.objects.filter(key='distance_rule').first()
            # Eğer yeni ayar yoksa, eski require_gps ayarını kontrol et
            if not distance_rule_setting:
                old_setting = SystemSetting.objects.filter(key='require_gps').first()
                if old_setting:
                    # Eski ayarı yeni isimle güncelle
                    old_setting.key = 'distance_rule'
                    old_setting.label = 'Mesafe Kuralı'
                    old_setting.description = 'Açık: Giriş mesafesi ve gezinme mesafesi kontrolü yapılır. Kapalı: Mesafe kontrolü yapılmaz, herhangi bir mesafeden ziyaret başlatılabilir.'
                    old_setting.save()
                    distance_rule_setting = old_setting
            
            distance_rule_enabled = True  # Varsayılan: açık
            if distance_rule_setting:
                try:
                    distance_rule_enabled = distance_rule_setting.value.lower() == 'true'
                except:
                    distance_rule_enabled = True
            
            # Eğer mesafe kuralı kapalıysa, direkt ziyareti başlat
            if not distance_rule_enabled:
                from datetime import datetime
                task.check_in_time = datetime.now()
                task.save()
                return JsonResponse({
                    'success': True, 
                    'message': 'Ziyaret başlatıldı. (Mesafe kuralı kapalı olduğu için mesafe kontrolü yapılmadı.)'
                })
            
            # 5. Admin Panelindeki Sınırı Çek (Mesafe kuralı açıksa)
            # Eğer ayar yoksa varsayılan 300 metre olsun
            setting = SystemSetting.objects.filter(key='visit_radius').first()
            try:
                limit = float(setting.value) if setting and setting.value else 300.0
            except (ValueError, TypeError):
                limit = 300.0

            # 6. KARAR ANI - Mesafe kontrolü (>= kullanarak sıkı kontrol)
            # Eğer mesafe limit'e eşit veya fazlaysa ziyaret başlatılmamalı
            if distance >= limit:
                # Mesafe UZAK veya EŞİT - Ziyaret başlatılmamalı
                distance_diff = int(distance - limit)
                return JsonResponse({
                    'success': False, 
                    'message': f"Ziyaret mesafesi uyarısı!\nTespit Edilen Mesafe: {int(distance)}m\nİzin Verilen: {int(limit)}m\nFark: {distance_diff}m fazla"
                })
            
            # Mesafe uygun (limit'ten küçük), ziyaret başlatıldı
            from datetime import datetime
            task.check_in_time = datetime.now()
            task.save()
            return JsonResponse({'success': True, 'message': 'Konum doğrulandı. Ziyaret başladı.'})

        except json.JSONDecodeError as e:
            import logging
            logger = logging.getLogger(__name__)
            try:
                logger.error(f"JSON decode hatası: {str(e)}, Body: {request.body}")
            except:
                pass
            return JsonResponse({'success': False, 'message': 'Geçersiz veri formatı. Lütfen tekrar deneyin.'})
        except ValueError as e:
            import logging
            logger = logging.getLogger(__name__)
            try:
                logger.error(f"ValueError: {str(e)}")
            except:
                pass
            return JsonResponse({'success': False, 'message': 'Koordinat bilgisi geçersiz. Lütfen konum iznini kontrol edin.'})
        except VisitTask.DoesNotExist:
            import logging
            logger = logging.getLogger(__name__)
            try:
                logger.error(f"VisitTask bulunamadı - Task ID: {task_id}")
            except:
                pass
            return JsonResponse({'success': False, 'message': 'Ziyaret görevi bulunamadı.'})
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            try:
                logger.error(f"Genel hata: {str(e)}", exc_info=True)
            except:
                pass
            return JsonResponse({'success': False, 'message': f'Bir hata oluştu: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Hatalı İstek - POST metodu bekleniyor.'})

# Zorunlu anketleri kontrol et
@csrf_exempt
@login_required
def check_required_surveys(request, task_id):
    """Ziyareti bitirmeden önce zorunlu anketlerin tamamlanıp tamamlanmadığını kontrol eder"""
    task = get_object_or_404(VisitTask, pk=task_id)
    if task.merch_code != request.user.username:
        return JsonResponse({'success': False, 'message': 'Yetkisiz.'}, status=403)
    
    # Bu görev için gösterilen tüm anketleri al - sadece bu tenant'a ait (temel yapı yok)
    task_tenant = task.customer.tenant if task.customer and task.customer.tenant else None
    surveys = Survey.objects.filter(is_active=True)
    
    # Tenant filtresi: Sadece bu görevin tenant'ına ait anketler (temel yapı yok)
    if task_tenant:
        surveys = surveys.filter(tenant=task_tenant)
    else:
        # Tenant yoksa, kullanıcının tenant'ını kullan
        user_tenant = getattr(request.user, 'tenant', None)
        if user_tenant:
            surveys = surveys.filter(tenant=user_tenant)
    
    from datetime import date
    today = date.today()
    surveys = surveys.filter(
        models.Q(start_date__isnull=True) | models.Q(start_date__lte=today)
    ).filter(
        models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
    )
    
    # Filtreleme (mobile_task_detail ile aynı mantık)
    user = request.user
    customer = task.customer
    filtered_surveys = []
    
    for survey in surveys:
        should_show = True
        
        if survey.filter_users.exists():
            if user not in survey.filter_users.all():
                should_show = False
        
        if survey.target_roles.exists():
            if not user.role or user.role not in survey.target_roles.all():
                should_show = False
        
        if survey.user_custom_filters:
            for field_slug, allowed_values in survey.user_custom_filters.items():
                if allowed_values:
                    user_value_str = user.extra_data.get(field_slug, '') if user.extra_data else ''
                    user_tags = [tag.strip() for tag in str(user_value_str).split(',') if tag.strip()]
                    if not any(tag in allowed_values for tag in user_tags):
                        should_show = False
                        break
        
        if survey.filter_customers.exists():
            if customer not in survey.filter_customers.all():
                should_show = False
        
        if survey.filter_caris.exists():
            if not customer.cari or customer.cari not in survey.filter_caris.all():
                should_show = False
        
        if survey.custom_filters:
            for field_slug, allowed_values in survey.custom_filters.items():
                if allowed_values:
                    customer_value = customer.extra_data.get(field_slug, '') if customer.extra_data else ''
                    if customer_value not in allowed_values:
                        should_show = False
                        break
        
        if should_show:
            filtered_surveys.append(survey)
    
    # Zorunlu soruları kontrol et
    missing_required = []
    for survey in filtered_surveys:
        # Bu anketin tüm zorunlu sorularını al
        all_questions = survey.questions.all()
        required_questions = [q for q in all_questions if q.required]
        
        # Bu görev için bu anketin sorularına verilen cevapları kontrol et
        for req_q in required_questions:
            answer = SurveyAnswer.objects.filter(task=task, question=req_q).first()
            if not answer or (not answer.answer_text and not answer.answer_photo):
                missing_required.append(survey)
                break  # Bu anket eksik, diğer sorularına bakmaya gerek yok
    
    return JsonResponse({
        'missing_required': [{'id': s.id, 'title': s.title} for s in missing_required],
        'all_completed': len(missing_required) == 0
    })

# Ziyareti bitir
@csrf_exempt
@login_required
def finish_visit(request, task_id):
    """Ziyareti tamamla - Sadece buton ile bitirilebilir, zorunlu formlar kontrol edilir"""
    task = get_object_or_404(VisitTask, pk=task_id)
    if task.merch_code != request.user.username:
        return JsonResponse({'success': False, 'message': 'Bu ziyareti bitirme yetkiniz yok.'}, status=403)
    
    # Zorunlu anketleri kontrol et - Tenant filtresi ekle
    task_tenant = task.customer.tenant if task.customer and task.customer.tenant else None
    surveys = Survey.objects.filter(is_active=True)
    
    # Tenant filtresi: Sadece bu görevin tenant'ına ait anketleri al
    if task_tenant:
        surveys = surveys.filter(tenant=task_tenant)
    else:
        # Tenant yoksa, kullanıcının tenant'ını kullan
        user_tenant = getattr(request.user, 'tenant', None)
        if user_tenant:
            surveys = surveys.filter(tenant=user_tenant)
    
    from datetime import date
    today = date.today()
    surveys = surveys.filter(
        models.Q(start_date__isnull=True) | models.Q(start_date__lte=today)
    ).filter(
        models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
    )
    
    user = request.user
    customer = task.customer
    filtered_surveys = []
    
    for survey in surveys:
        should_show = True
        
        if survey.filter_users.exists():
            if user not in survey.filter_users.all():
                should_show = False
        
        if survey.target_roles.exists():
            if not user.role or user.role not in survey.target_roles.all():
                should_show = False
        
        if survey.user_custom_filters:
            for field_slug, allowed_values in survey.user_custom_filters.items():
                if allowed_values:
                    user_value_str = user.extra_data.get(field_slug, '') if user.extra_data else ''
                    user_tags = [tag.strip() for tag in str(user_value_str).split(',') if tag.strip()]
                    if not any(tag in allowed_values for tag in user_tags):
                        should_show = False
                        break
        
        if survey.filter_customers.exists():
            if customer not in survey.filter_customers.all():
                should_show = False
        
        if survey.filter_caris.exists():
            if not customer.cari or customer.cari not in survey.filter_caris.all():
                should_show = False
        
        if survey.custom_filters:
            for field_slug, allowed_values in survey.custom_filters.items():
                if allowed_values:
                    customer_value = customer.extra_data.get(field_slug, '') if customer.extra_data else ''
                    if customer_value not in allowed_values:
                        should_show = False
                        break
        
        if should_show:
            filtered_surveys.append(survey)
    
    # Zorunlu soruları kontrol et
    missing_required = []
    for survey in filtered_surveys:
        all_questions = survey.questions.all()
        required_questions = [q for q in all_questions if q.required]
        
        for req_q in required_questions:
            answer = SurveyAnswer.objects.filter(task=task, question=req_q).first()
            if not answer or (not answer.answer_text and not answer.answer_photo):
                missing_required.append(survey)
                break
    
    if missing_required:
        return JsonResponse({
            'success': False,
            'message': 'Zorunlu anketler tamamlanmadan ziyaret bitirilemez.',
            'missing_required': [{'id': s.id, 'title': s.title} for s in missing_required]
        })
    
    # Tüm kontroller geçildi, ziyareti bitir
    from datetime import datetime
    task.status = 'completed'
    task.check_out_time = datetime.now()
    task.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Ziyaret başarıyla tamamlandı.'
    })

# Gezinme sınırını getir
@csrf_exempt
def get_wander_radius(request):
    """Gezinme sınırı ayarını döndürür"""
    setting = SystemSetting.objects.filter(key='wander_radius').first()
    wander_radius = float(setting.value) if setting else 500.0  # Varsayılan 500m
    return JsonResponse({'wander_radius': wander_radius})

@csrf_exempt
@login_required
def get_data_sync_interval(request):
    """Veri paylaşma süresini (dakika) döndürür"""
    setting = SystemSetting.objects.filter(key='data_sync_interval_minutes').first()
    interval = int(setting.value) if setting else 15  # Varsayılan 15 dakika
    return JsonResponse({'interval_minutes': interval})

@login_required
def mobile_sync_pending_data(request):
    """Bekleyen form verilerini hemen gönder"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Sadece POST istekleri kabul edilir.'}, status=405)
    
    try:
        # localStorage'dan bekleyen form verilerini al (JavaScript tarafından gönderilecek)
        import json
        data = json.loads(request.body) if request.body else {}
        pending_forms = data.get('pending_forms', [])
        
        success_count = 0
        error_count = 0
        
        for form_data in pending_forms:
            try:
                task_id = form_data.get('task_id')
                survey_id = form_data.get('survey_id')
                answers = form_data.get('answers', {})
                
                task = VisitTask.objects.get(pk=task_id, merch_code=request.user.username)
                survey = Survey.objects.get(pk=survey_id)
                
                # Form verilerini kaydet
                for q_id, answer_data in answers.items():
                    question = Question.objects.get(pk=q_id, survey=survey)
                    text_val = answer_data.get('text')
                    photo_b64 = answer_data.get('photo_base64')
                    
                    # Eski cevabı sil
                    SurveyAnswer.objects.filter(task=task, question=question).delete()
                    
                    # Fotoğraf varsa işle
                    photo_val = None
                    if photo_b64 and photo_b64.startswith('data:image/'):
                        try:
                            header, b64data = photo_b64.split(',', 1)
                            ext = 'jpg'
                            if 'image/png' in header:
                                ext = 'png'
                            elif 'image/webp' in header:
                                ext = 'webp'
                            
                            import base64
                            decoded = base64.b64decode(b64data)
                            photo_val = ContentFile(decoded, name=f"survey_{task_id}_{q_id}.{ext}")
                        except Exception:
                            pass
                    
                    # Yeni cevabı kaydet
                    if text_val or photo_val:
                        SurveyAnswer.objects.create(
                            task=task,
                            question=question,
                            answer_text=text_val,
                            answer_photo=photo_val
                        )
                
                success_count += 1
            except Exception as e:
                error_count += 1
                print(f"Form gönderim hatası: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'{success_count} form gönderildi.',
            'success_count': success_count,
            'error_count': error_count
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Hata: {str(e)}'}, status=500)

@csrf_exempt
@login_required
def get_distance_rule(request):
    """Mesafe kuralı ayarını döndürür"""
    setting = SystemSetting.objects.filter(key='distance_rule').first()
    # Eğer eski require_gps ayarı varsa, onu distance_rule olarak kullan
    if not setting:
        old_setting = SystemSetting.objects.filter(key='require_gps').first()
        if old_setting:
            # Eski ayarı yeni isimle güncelle
            old_setting.key = 'distance_rule'
            old_setting.label = 'Mesafe Kuralı'
            old_setting.description = 'Açık: Giriş mesafesi ve gezinme mesafesi kontrolü yapılır. Kapalı: Mesafe kontrolü yapılmaz, herhangi bir mesafeden ziyaret başlatılabilir.'
            old_setting.save()
            setting = old_setting
    
    distance_rule = True  # Varsayılan: açık
    if setting:
        try:
            distance_rule = setting.value.lower() == 'true'
        except:
            distance_rule = True
    
    return JsonResponse({'distance_rule': distance_rule})

@csrf_exempt
@login_required
def check_visit_status(request, task_id):
    """Ziyaretin başlatılıp başlatılmadığını kontrol eder"""
    task = get_object_or_404(VisitTask, pk=task_id)
    if task.merch_code != request.user.username:
        return JsonResponse({'success': False, 'message': 'Yetkisiz.'}, status=403)
    
    # Ziyaret başlatılmışsa check_in_time dolu olur, bitirilmişse check_out_time dolu olur
    is_started = task.check_in_time is not None
    is_finished = task.check_out_time is not None and task.status == 'completed'
    
    return JsonResponse({
        'is_started': is_started and not is_finished,
        'is_finished': is_finished,
        'check_in_time': task.check_in_time.isoformat() if task.check_in_time else None,
        'check_out_time': task.check_out_time.isoformat() if task.check_out_time else None
    })
    
    return JsonResponse({'distance_rule': distance_rule})


@login_required
def mobile_team_member(request, merch_code: str):
    """
    Supervisor view: show today's tasks for a subordinate merch_code.
    """
    user = request.user
    scope = get_hierarchy_scope_for_user(user, include_self=True)
    if scope.usernames and merch_code not in scope.usernames:
        return HttpResponseForbidden("Bu personeli görüntüleme yetkiniz yok.")

    today = date.today()
    tasks = (
        VisitTask.objects.filter(merch_code=merch_code, planned_date=today)
        .select_related("customer")
        .order_by("status", "customer__name")
    )
    total = tasks.count()
    completed = tasks.filter(status="completed").count()
    ctx = {
        "today": today,
        "merch_code": merch_code,
        "tasks": tasks,
        "total": total,
        "completed": completed,
    }
    return render(request, "mobile/team_member_tasks.html", ctx)

# --- OTOMATIK MAIL YÖNETİMİ ---
from apps.core.models import AutomatedEmail
from apps.core.tenant_utils import set_tenant_on_save
from django.urls import reverse

def _resolve_tenant_for_automated_email(request):
    """Request'ten tenant'ı çözümler (root admin veya normal kullanıcı için)"""
    from apps.core.tenant_utils import get_current_tenant
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        tenant = get_current_tenant(request)
    return tenant

def _get_available_reports_for_tenant(tenant):
    """
    Tenant için mevcut raporları hiyerarşik yapıda döndürür
    Sadece "Raporlar" bölümündeki raporlar: Detaylı Ziyaret, Günlük Özet, Anket Raporları
    """
    from django.contrib.contenttypes.models import ContentType
    from apps.field_operations.models import ReportRecord
    from apps.forms.models import Survey
    
    reports_list = []
    
    # 1. Detaylı Ziyaret Raporu
    reports_list.append({
        'key': 'visit_detail',
        'name': 'Detaylı Ziyaret Raporu',
        'description': 'Tüm ziyaret detaylarını içeren kapsamlı rapor',
        'type': 'single'
    })
    
    # 2. Günlük Özet Raporu
    reports_list.append({
        'key': 'daily_summary',
        'name': 'Günlük Özet Raporu',
        'description': 'Günlük ziyaret özet raporu',
        'type': 'single'
    })
    
    # 3. Anket Raporları grubu (ReportRecord sisteminden)
    ct = ContentType.objects.get_for_model(Survey)
    survey_report_records = ReportRecord.objects.filter(
        tenant=tenant,
        report_type='survey',
        content_type=ct,
        deleted_at__isnull=True
    ).order_by('title')
    
    if survey_report_records.exists():
        survey_reports = []
        for record in survey_report_records:
            try:
                survey = Survey.objects.get(id=record.object_id, tenant=tenant)
                survey_reports.append({
                    'key': f'survey_{survey.id}',
                    'name': survey.title,
                    'description': f'Anket: {survey.title}',
                    'type': 'single'
                })
            except Survey.DoesNotExist:
                continue
        
        if survey_reports:
            reports_list.append({
                'key': 'survey_group',
                'name': 'Anket Raporları',
                'description': 'Anket sonuçlarını içeren raporlar',
                'type': 'group',
                'children': survey_reports
            })
    
    return reports_list

@login_required
def automated_email_list(request):
    is_admin_panel = getattr(request, 'is_admin_panel', False)
    is_root = is_root_admin(request.user)
    
    # selected_tenant_id'yi başlangıçta None olarak tanımla
    selected_tenant_id = None

    if is_root and is_admin_panel:
        selected_tenant_id = request.GET.get('tenant_id')
        if selected_tenant_id:
            tenant = get_object_or_404(Tenant, id=selected_tenant_id, is_active=True)
        else:
            tenant = None
        all_tenants = Tenant.objects.filter(is_active=True).order_by('name')
    else:
        tenant = _resolve_tenant_for_automated_email(request)
        all_tenants = []
    
    if not tenant and not is_root:
        messages.error(request, "Firma bilgisi bulunamadı.")
        return redirect("home")
    elif not tenant and is_root and not selected_tenant_id:
        automated_emails = AutomatedEmail.objects.none()
        available_reports = []
    else:
        automated_emails = AutomatedEmail.objects.filter(tenant=tenant).order_by("-created_at")
        available_reports = _get_available_reports_for_tenant(tenant)

    context = {
        "automated_emails": automated_emails,
        "available_reports": available_reports,
        "tenant": tenant,
        "is_root_admin": is_root,
        "all_tenants": all_tenants,
        "selected_tenant_id": int(selected_tenant_id) if selected_tenant_id else None,
    }
    return render(request, "apps/core/automated_email_list.html", context)

@login_required
def automated_email_create(request):
    is_root = is_root_admin(request.user)
    is_admin_panel = getattr(request, 'is_admin_panel', False)

    if is_root and is_admin_panel:
        all_tenants = Tenant.objects.filter(is_active=True).order_by('name')
        selected_tenant_id = request.POST.get('tenant_id') if request.method == 'POST' else request.GET.get('tenant_id')
        if selected_tenant_id:
            tenant = get_object_or_404(Tenant, id=selected_tenant_id, is_active=True)
        else:
            tenant = None
    else:
        tenant = _resolve_tenant_for_automated_email(request)
        all_tenants = []
        selected_tenant_id = None

    if not tenant:
        messages.error(request, "Lütfen önce bir firma seçin.")
        if is_root and is_admin_panel:
            return redirect('automated_email_list')
        return redirect("home")

    available_reports = _get_available_reports_for_tenant(tenant)

    if request.method == "POST":
        to_email = request.POST.get('to_email', '').strip()
        cc_email = request.POST.get('cc_email', '').strip()
        subject = request.POST.get('subject', '').strip()
        body = request.POST.get('body', '').strip()
        
        selected_reports = {}
        # Template'de name="selected_reports" olarak gönderiliyor, value attribute'undan rapor key'ini al
        selected_report_keys = request.POST.getlist('selected_reports')
        for report_key in selected_report_keys:
            if report_key:  # Boş değilse
                selected_reports[report_key] = True
        
        merge_reports = request.POST.get('merge_reports') == 'on'
        
        # Tarih alanları - boş string kontrolü (Django DateField boş string kabul etmez)
        report_start_date = request.POST.get('report_start_date', '').strip() or None
        report_end_date = request.POST.get('report_end_date', '').strip() or None
        send_start_date = request.POST.get('send_start_date', '').strip() or None
        send_end_date = request.POST.get('send_end_date', '').strip() or None
        period = request.POST.get('period')
        day_option = request.POST.get('day_option') or None
        send_time = request.POST.get('send_time')
        
        if not to_email or not subject or not body:
            messages.error(request, "Kime, Konu ve Mail İçeriği alanları zorunludur.")
            return render(request, "apps/core/automated_email_form.html", {
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
                "selected_tenant_id": selected_tenant_id,
            })
        
        if not selected_reports:
            messages.error(request, "En az bir rapor seçmelisiniz.")
            return render(request, "apps/core/automated_email_form.html", {
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
                "selected_tenant_id": selected_tenant_id,
            })
        
        # Tarih validasyonu - zorunlu alanlar
        if not report_start_date:
            messages.error(request, "Rapor Başlangıç Tarihi zorunludur.")
            return render(request, "apps/core/automated_email_form.html", {
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
                "selected_tenant_id": selected_tenant_id,
            })
        
        if not report_end_date:
            messages.error(request, "Rapor Bitiş Tarihi zorunludur.")
            return render(request, "apps/core/automated_email_form.html", {
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
                "selected_tenant_id": selected_tenant_id,
            })
        
        if not send_start_date:
            messages.error(request, "Gönderim Başlangıç Tarihi zorunludur.")
            return render(request, "apps/core/automated_email_form.html", {
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
                "selected_tenant_id": selected_tenant_id,
            })
        
        if is_root and is_admin_panel:
            tenant_id_from_post = request.POST.get('tenant_id')
            if not tenant_id_from_post:
                messages.error(request, "Lütfen bir firma seçin.")
                return render(request, "apps/core/automated_email_form.html", {
                    "available_reports": available_reports,
                    "tenant": tenant,
                    "is_root_admin": is_root,
                    "all_tenants": all_tenants,
                    "selected_tenant_id": selected_tenant_id,
                })
            tenant_to_save = get_object_or_404(Tenant, id=tenant_id_from_post, is_active=True)
        else:
            tenant_to_save = tenant
        
        is_active = request.POST.get('is_active') == 'on'
        
        automated = AutomatedEmail.objects.create(
            tenant=tenant_to_save,
            to_email=to_email,
            cc_email=cc_email if cc_email else None,
            subject=subject,
            body=body,
            selected_reports=selected_reports,
            merge_reports=merge_reports,
            report_start_date=report_start_date,
            report_end_date=report_end_date,
            send_start_date=send_start_date,
            send_end_date=send_end_date if send_end_date else None,
            period=period,
            day_option=day_option,
            send_time=send_time,
            is_active=is_active,
            created_by=request.user,
        )
        messages.success(request, "Otomatik mail başarıyla oluşturuldu.")
        return redirect("automated_email_list")

    context = {
        "available_reports": available_reports,
        "tenant": tenant,
        "is_root_admin": is_root,
        "all_tenants": all_tenants,
        "selected_tenant_id": int(selected_tenant_id) if selected_tenant_id else None,
    }
    return render(request, "apps/core/automated_email_form.html", context)

@login_required
def automated_email_edit(request, pk):
    automated_email = get_object_or_404(AutomatedEmail, pk=pk)
    
    is_root = is_root_admin(request.user)
    if not is_root and automated_email.tenant != _resolve_tenant_for_automated_email(request):
        messages.error(request, "Bu otomatik maili düzenleme yetkiniz yok.")
        return redirect("automated_email_list")
    
    tenant = automated_email.tenant
    available_reports = _get_available_reports_for_tenant(tenant)
    is_admin_panel = getattr(request, 'is_admin_panel', False)
    all_tenants = Tenant.objects.filter(is_active=True).order_by('name') if is_root and is_admin_panel else []

    if request.method == "POST":
        automated_email.to_email = request.POST.get('to_email', '').strip()
        automated_email.cc_email = request.POST.get('cc_email', '').strip()
        automated_email.subject = request.POST.get('subject', '').strip()
        automated_email.body = request.POST.get('body', '').strip()
        
        selected_reports = {}
        # Template'de name="selected_reports" olarak gönderiliyor, value attribute'undan rapor key'ini al
        selected_report_keys = request.POST.getlist('selected_reports')
        for report_key in selected_report_keys:
            if report_key:  # Boş değilse
                selected_reports[report_key] = True
        
        automated_email.selected_reports = selected_reports
        automated_email.merge_reports = request.POST.get('merge_reports') == 'on'
        
        # Tarih alanları - boş string kontrolü (Django DateField boş string kabul etmez)
        report_start_date = request.POST.get('report_start_date', '').strip() or None
        report_end_date = request.POST.get('report_end_date', '').strip() or None
        send_start_date = request.POST.get('send_start_date', '').strip() or None
        send_end_date = request.POST.get('send_end_date', '').strip() or None
        
        # Zorunlu alanlar için validation
        if not report_start_date:
            messages.error(request, "Rapor Başlangıç Tarihi zorunludur.")
            return render(request, "apps/core/automated_email_form.html", {
                "automated_email": automated_email,
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
            })
        
        if not report_end_date:
            messages.error(request, "Rapor Bitiş Tarihi zorunludur.")
            return render(request, "apps/core/automated_email_form.html", {
                "automated_email": automated_email,
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
            })
        
        if not send_start_date:
            messages.error(request, "Gönderim Başlangıç Tarihi zorunludur.")
            return render(request, "apps/core/automated_email_form.html", {
                "automated_email": automated_email,
                "available_reports": available_reports,
                "tenant": tenant,
                "is_root_admin": is_root,
                "all_tenants": all_tenants,
            })
        
        automated_email.report_start_date = report_start_date
        automated_email.report_end_date = report_end_date
        automated_email.send_start_date = send_start_date
        automated_email.send_end_date = send_end_date
        automated_email.period = request.POST.get('period')
        automated_email.day_option = request.POST.get('day_option') or None
        automated_email.send_time = request.POST.get('send_time')
        automated_email.is_active = request.POST.get('is_active') == 'on'
        
        automated_email.save()
        messages.success(request, "Otomatik mail başarıyla güncellendi.")
        return redirect("automated_email_list")

    context = {
        "automated_email": automated_email,
        "available_reports": available_reports,
        "tenant": tenant,
        "is_root_admin": is_root,
        "all_tenants": all_tenants,
    }
    return render(request, "apps/core/automated_email_form.html", context)

def _generate_report_for_automated_email(tenant, report_key, start_date, end_date):
    """
    Otomatik mail için rapor oluşturur ve Excel bytes döndürür
    Returns: (filename, excel_bytes)
    """
    from datetime import date
    from apps.field_operations.models import VisitTask
    from apps.field_operations.views import (
        _visit_detail_report_columns, _task_to_report_value,
        _build_user_and_hierarchy_maps
    )
    from apps.core.excel_utils import xlsx_from_rows
    from django.db.models import Count, Min, Max, Q
    from django.utils import timezone
    
    if report_key == 'visit_detail':
        # Ziyaret Detay Raporu
        cols, label_by_key = _visit_detail_report_columns()
        selected_cols = [
            "answer_id", "customer_code", "customer_name",
            "visit_start_date", "visit_start_time", "visit_end_date",
            "visit_end_time", "visit_duration_min"
        ]
        
        qs = VisitTask.objects.select_related("customer", "customer__cari").filter(tenant=tenant)
        qs = qs.exclude(check_in_time__isnull=True)
        if start_date and end_date:
            qs = qs.filter(check_in_time__date__gte=start_date, check_in_time__date__lte=end_date)
        qs = qs.order_by("-check_in_time", "-id")
        
        usernames = list(qs.values_list("merch_code", flat=True).distinct())
        user_fullname_by_username, hierarchy_parent_by_username = _build_user_and_hierarchy_maps(usernames)
        
        rows = []
        for t in qs:
            rows.append({
                k: _task_to_report_value(
                    t, k,
                    user_fullname_by_username=user_fullname_by_username,
                    hierarchy_parent_by_username=hierarchy_parent_by_username,
                )
                for k in selected_cols
            })
        
        excel_bytes = xlsx_from_rows(rows, sheet_name="Ziyaret Detay Raporu", header_order=selected_cols, label_by_key=label_by_key)
        return ("Ziyaret_Detay_Raporu.xlsx", excel_bytes)
    
    elif report_key == 'daily_summary':
        # Günlük Özet Raporu
        qs = VisitTask.objects.select_related("customer").filter(tenant=tenant, planned_date__gte=start_date, planned_date__lte=end_date)
        
        stats_qs = (
            qs.values("merch_code", "planned_date")
            .annotate(
                planned=Count("id"),
                completed=Count("id", filter=Q(status="completed")),
                store_count=Count("customer_id", distinct=True),
            )
            .order_by("planned_date", "merch_code")
        )
        
        all_merchs = [r["merch_code"] for r in stats_qs if r.get("merch_code")]
        full_by_username, _ = _build_user_and_hierarchy_maps(all_merchs)
        
        rows = []
        for r in stats_qs:
            merch = r.get("merch_code") or ""
            planned_date = r.get("planned_date")
            rows.append({
                "Tarih": planned_date.strftime("%d.%m.%Y") if planned_date else "",
                "Personel": merch,
                "Personel Adı": full_by_username.get(merch, "") or merch,
                "Planlanan": int(r.get("planned") or 0),
                "Tamamlanan": int(r.get("completed") or 0),
                "Mağaza Sayısı": int(r.get("store_count") or 0),
            })
        
        excel_bytes = xlsx_from_rows(rows, sheet_name="Günlük Özet")
        return ("Gunluk_Ozet_Raporu.xlsx", excel_bytes)
    
    # Müşteri Listesi ve Rota Planı kaldırıldı - Sadece "Raporlar" bölümündeki raporlar gönderilir
    
    elif report_key.startswith('survey_'):
        # Anket Raporu (ReportRecord sisteminden)
        from django.contrib.contenttypes.models import ContentType
        from apps.forms.models import Survey, SurveyAnswer
        
        try:
            survey_id = int(report_key.replace('survey_', ''))
            survey = Survey.objects.get(id=survey_id, tenant=tenant)
        except (Survey.DoesNotExist, ValueError) as e:
            print(f"[ERROR] Survey not found or invalid ID: {report_key}, error: {str(e)}")
            return None
        
        # Survey raporu için mevcut export fonksiyonunu kullan
        try:
            from apps.field_operations.views import (
                _survey_report_columns, _task_to_survey_report_value,
                _build_user_and_hierarchy_maps
            )
            
            print(f"[DEBUG] Generating survey report for survey_id: {survey_id}, title: {survey.title}")
            
            cols, label_by_key = _survey_report_columns(survey)
            qs = (
                VisitTask.objects.select_related("customer")
                .filter(tenant=tenant, answers__question__survey=survey)
                .exclude(check_in_time__isnull=True)
                .distinct()
            )
            
            print(f"[DEBUG] Base queryset count: {qs.count()}")
            
            if start_date and end_date:
                qs = qs.filter(check_in_time__date__gte=start_date, check_in_time__date__lte=end_date)
                print(f"[DEBUG] After date filter: {qs.count()}")
            
            qs = qs.order_by("-check_in_time", "-id")
            
            task_ids = list(qs.values_list("id", flat=True))
            print(f"[DEBUG] Task IDs found: {len(task_ids)}")
            
            if not task_ids:
                print(f"[WARNING] No tasks found for survey {survey_id} in date range {start_date} to {end_date}")
                # Boş rapor oluştur - ama tüm kolonları ekle
                rows = []
                selected_cols = ["answer_id", "customer_code", "customer_name", "personel"]
                q_cols = [c["key"] for c in cols if c["key"].startswith("q_")]
                selected_cols.extend(q_cols)
                
                # Boş rapor için Excel oluştur (başlıklar doğru görünsün)
                excel_bytes = xlsx_from_rows(
                    rows, 
                    sheet_name=survey.title[:31], 
                    header_order=selected_cols,
                    label_by_key=label_by_key  # Başlıkları label'larla değiştir
                )
                print(f"[DEBUG] Empty survey report Excel created with columns: {selected_cols}")
                return (f"Anket_{survey.title[:20]}.xlsx", excel_bytes)
            else:
                answers = (
                    SurveyAnswer.objects.select_related("question")
                    .filter(task_id__in=task_ids, question__survey=survey)
                    .order_by("task_id", "question_id", "id")
                )
                
                print(f"[DEBUG] Answers found: {answers.count()}")
                
                bucket = {}
                for a in answers:
                    tid = a.task_id
                    qid = a.question_id
                    val = ""
                    if a.answer_text:
                        val = str(a.answer_text).strip()
                    elif a.answer_photo:
                        try:
                            val = a.answer_photo.url
                        except:
                            val = str(a.answer_photo)
                    if val:
                        bucket.setdefault((tid, qid), []).append(val)
                answers_map = {k: " | ".join(v) for k, v in bucket.items()}
                
                usernames = list(qs.values_list("merch_code", flat=True).distinct())
                user_fullname_by_username, hierarchy_parent_by_username = _build_user_and_hierarchy_maps(usernames)
                
                selected_cols = ["answer_id", "customer_code", "customer_name", "personel"]
                q_cols = [c["key"] for c in cols if c["key"].startswith("q_")]
                selected_cols.extend(q_cols)
                
                rows = []
                for t in qs:
                    row = {
                        k: _task_to_survey_report_value(
                            t, k,
                            answers_map=answers_map,
                            user_fullname_by_username=user_fullname_by_username,
                            hierarchy_parent_by_username=hierarchy_parent_by_username,
                        )
                        for k in selected_cols
                    }
                    rows.append(row)
            
            excel_bytes = xlsx_from_rows(
                rows, 
                sheet_name=survey.title[:31], 
                header_order=selected_cols,
                label_by_key=label_by_key  # Başlıkları label'larla değiştir
            )
            print(f"[DEBUG] Survey report Excel created: {len(excel_bytes)} bytes")
            print(f"[DEBUG] Columns: {selected_cols}")
            print(f"[DEBUG] Labels: {[label_by_key.get(k, k) for k in selected_cols]}")
            return (f"Anket_{survey.title[:20]}.xlsx", excel_bytes)
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"[ERROR] Survey report generation error: {error_trace}")
            raise  # Exception'ı yukarı fırlat ki ana fonksiyonda yakalansın
    
    return None

def _update_email_settings_from_db_for_tenant(tenant):
    """Email ayarlarını veritabanından yükleyip Django settings'e ekle (TENANT-SPECIFIC)"""
    try:
        from django.conf import settings
        from apps.core.models import SystemSetting
        
        # Veritabanından email ayarlarını oku (tenant'a özel)
        email_settings = {}
        settings_qs = SystemSetting.objects.filter(category='email', tenant=tenant)
        
        # Tenant için email ayarları yoksa, Rotexia şablonunu kullan
        if not settings_qs.exists():
            print(f"[WARNING] Tenant ({tenant.name}) için email ayarları bulunamadı, Rotexia şablonu kullanılıyor.")
            settings_qs = SystemSetting.objects.filter(category='email', tenant__isnull=True)
        
        for setting in settings_qs:
            key = setting.key.replace('email_', '')
            email_settings[key] = setting.value
        
        # Veritabanındaki ayarları her zaman Django settings'e yükle (tenant'a özel ayarlar öncelikli)
        if email_settings.get('host'):
            settings.EMAIL_HOST = email_settings.get('host', '')
        
        if email_settings.get('host_user'):
            settings.EMAIL_HOST_USER = email_settings.get('host_user', '')
        
        if email_settings.get('host_password'):
            # Password'ü trim et (baş/son boşlukları temizle)
            password = email_settings.get('host_password', '').strip()
            settings.EMAIL_HOST_PASSWORD = password
        
        if email_settings.get('port'):
            try:
                settings.EMAIL_PORT = int(email_settings.get('port', '587'))
            except:
                settings.EMAIL_PORT = 587
        
        # TLS/SSL ayarları - Outlook/Hotmail için özel kontrol
        use_tls = email_settings.get('use_tls', 'True').lower() == 'true'
        
        # Outlook/Hotmail için özel ayarlar
        if 'outlook.com' in settings.EMAIL_HOST or 'office365.com' in settings.EMAIL_HOST:
            settings.EMAIL_USE_TLS = True  # STARTTLS kullan (587 port için)
            settings.EMAIL_USE_SSL = False  # SSL kullanma
            if settings.EMAIL_PORT == 465:
                settings.EMAIL_PORT = 587  # Outlook için 587 port kullan (STARTTLS için)
        else:
            settings.EMAIL_USE_TLS = use_tls
            settings.EMAIL_USE_SSL = False
        
        # DEFAULT_FROM_EMAIL ayarla - her zaman veritabanından alınan değeri kullan
        default_from_email_value = email_settings.get('default_from_email', '').strip()
        if default_from_email_value:
            settings.DEFAULT_FROM_EMAIL = default_from_email_value
        else:
            # Veritabanında yoksa veya boşsa varsayılan değer kullan
            settings.DEFAULT_FROM_EMAIL = 'noreply@rotexia.com'
        
        # Debug: Ayarları logla
        print(f"[DEBUG] Email settings loaded for tenant: {tenant.name}")
        print(f"[DEBUG]   EMAIL_HOST: {getattr(settings, 'EMAIL_HOST', 'Not set')}")
        print(f"[DEBUG]   EMAIL_PORT: {getattr(settings, 'EMAIL_PORT', 'Not set')}")
        print(f"[DEBUG]   EMAIL_USE_TLS: {getattr(settings, 'EMAIL_USE_TLS', 'Not set')}")
        print(f"[DEBUG]   EMAIL_HOST_USER: {getattr(settings, 'EMAIL_HOST_USER', 'Not set')}")
        print(f"[DEBUG]   DEFAULT_FROM_EMAIL: {getattr(settings, 'DEFAULT_FROM_EMAIL', 'Not set')}")
        print(f"[DEBUG]   EMAIL_HOST_PASSWORD: {'***' + getattr(settings, 'EMAIL_HOST_PASSWORD', '')[-4:] if getattr(settings, 'EMAIL_HOST_PASSWORD', '') else 'Not set'}")
        
        # Email backend'i belirle
        if settings.EMAIL_HOST and settings.EMAIL_HOST_USER and settings.EMAIL_HOST_PASSWORD:
            settings.EMAIL_BACKEND = 'django.core.mail.backends.smtp.EmailBackend'
            print(f"[DEBUG] Using SMTP backend")
        else:
            settings.EMAIL_BACKEND = 'django.core.mail.backends.console.EmailBackend'
            print(f"[WARNING] Using console backend - missing settings")
        
    except Exception as e:
        print(f"[WARNING] Email settings from DB could not be loaded: {str(e)}")

def _merge_excel_reports(reports_list):
    """
    Birden fazla Excel raporunu tek bir Excel dosyasında birleştirir (her biri ayrı sheet)
    reports_list: [(filename, excel_bytes), ...]
    Returns: merged_excel_bytes
    """
    from openpyxl import load_workbook, Workbook
    from io import BytesIO
    
    if not reports_list:
        return None
    
    if len(reports_list) == 1:
        return reports_list[0][1]
    
    wb_merged = Workbook()
    wb_merged.remove(wb_merged.active)  # Varsayılan sheet'i kaldır
    
    for filename, excel_bytes in reports_list:
        wb = load_workbook(BytesIO(excel_bytes))
        for ws in wb.worksheets:
            ws_copy = wb_merged.create_sheet(title=ws.title[:31])
            for row in ws.iter_rows():
                ws_copy.append([cell.value for cell in row])
    
    out = BytesIO()
    wb_merged.save(out)
    return out.getvalue()

def _send_automated_email(automated_email, force=False):
    """
    Otomatik maili gönderir
    force=True ise zamanlama kontrolü yapmaz
    """
    from django.utils import timezone
    from django.core.mail import EmailMessage
    from datetime import datetime, date, timedelta
    import pytz
    
    # Türkiye saat dilimine göre kontrol et (UTC+3)
    turkey_tz = pytz.timezone('Europe/Istanbul')
    now = timezone.now()  # UTC zaman (last_sent_at için)
    now_turkey = now.astimezone(turkey_tz)  # Türkiye saati (zamanlama kontrolü için)
    today = now_turkey.date()
    current_time_turkey = now_turkey.time()
    
    # Zamanlama kontrolü (force=False ise)
    if not force:
        # Gönderim başlangıç tarihi kontrolü
        if automated_email.send_start_date > today:
            return False, "Gönderim başlangıç tarihi henüz gelmedi"
        
        # Gönderim bitiş tarihi kontrolü
        if automated_email.send_end_date and automated_email.send_end_date < today:
            return False, "Gönderim bitiş tarihi geçti"
        
        # Periyot ve gün kontrolü
        if automated_email.period == 'daily':
            if automated_email.day_option != 'every_day':
                return False, "Günlük periyot için sadece 'Her Gün' seçilebilir"
            # Günde bir kez kontrolü - bugün zaten gönderilmişse tekrar gönderme
            if automated_email.last_sent_at:
                last_sent_date_turkey = automated_email.last_sent_at.astimezone(turkey_tz).date()
                if last_sent_date_turkey == today:
                    return False, f"Bugün zaten gönderildi (Son gönderim: {automated_email.last_sent_at.astimezone(turkey_tz).strftime('%d.%m.%Y %H:%M')} TSİ)"
            # Saat kontrolü - Türkiye saatine göre kontrol et
            send_time = automated_email.send_time
            if send_time:
                from datetime import datetime
                send_datetime = datetime.combine(today, send_time)
                now_datetime = datetime.combine(today, current_time_turkey)
                
                print(f"[DEBUG] Gönderim kontrolü (Türkiye Saati) - Şimdi: {now_datetime.strftime('%Y-%m-%d %H:%M')}, Gönderim saati: {send_datetime.strftime('%Y-%m-%d %H:%M')}")
                
                if now_datetime < send_datetime:
                    return False, f"Gönderim saati henüz gelmedi (Gönderim saati: {send_time.strftime('%H:%M')} TSİ, Şimdi: {current_time_turkey.strftime('%H:%M')} TSİ)"
        
        elif automated_email.period == 'weekly':
            if today.weekday() != 0:  # Pazartesi = 0
                return False, "Haftalık periyot sadece pazartesi günleri gönderilir"
            if automated_email.day_option != 'monday':
                return False, "Haftalık periyot için 'Her Pazartesi' seçilmelidir"
            send_time = automated_email.send_time
            if send_time:
                from datetime import datetime
                send_datetime = datetime.combine(today, send_time)
                now_datetime = datetime.combine(today, current_time_turkey)
                
                if now_datetime < send_datetime:
                    return False, f"Gönderim saati henüz gelmedi (Gönderim saati: {send_time.strftime('%H:%M')} TSİ, Şimdi: {current_time_turkey.strftime('%H:%M')} TSİ)"
        
        elif automated_email.period == 'monthly':
            if today.day != 1:
                return False, "Aylık periyot sadece ayın ilk günü gönderilir"
            if automated_email.day_option != 'first_of_month':
                return False, "Aylık periyot için 'Her Ayın İlk Günü' seçilmelidir"
            send_time = automated_email.send_time
            if send_time:
                from datetime import datetime
                send_datetime = datetime.combine(today, send_time)
                now_datetime = datetime.combine(today, current_time_turkey)
                
                if now_datetime < send_datetime:
                    return False, f"Gönderim saati henüz gelmedi (Gönderim saati: {send_time.strftime('%H:%M')} TSİ, Şimdi: {current_time_turkey.strftime('%H:%M')} TSİ)"
        
        # "Her Gün" periyotunda günde bir kez kontrolü var
        # Bugün zaten gönderilmişse, yarın bekleyecek
    
    # Raporları oluştur
    selected_reports = automated_email.selected_reports or {}
    active_reports = [k for k, v in selected_reports.items() if v]
    
    print(f"[DEBUG] Active reports: {active_reports}")
    print(f"[DEBUG] Selected reports dict: {selected_reports}")
    
    if not active_reports:
        return False, "Hiç rapor seçilmemiş. Lütfen 'Raporlar' bölümünden en az bir rapor seçin."
    
    reports_list = []
    errors = []
    
    for report_key in active_reports:
        
        try:
            print(f"[DEBUG] Generating report: {report_key}")
            result = _generate_report_for_automated_email(
                automated_email.tenant,
                report_key,
                automated_email.report_start_date,
                automated_email.report_end_date
            )
            if result:
                reports_list.append(result)
                print(f"[DEBUG] Report generated successfully: {report_key}")
            else:
                errors.append(f"Rapor oluşturulamadı: {report_key}")
                print(f"[DEBUG] Report generation returned None: {report_key}")
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            error_msg = f"Rapor oluşturma hatası ({report_key}): {str(e)}"
            errors.append(error_msg)
            print(f"[ERROR] Report generation error ({report_key}): {error_trace}")
            continue
    
    if not reports_list:
        error_message = "Hiç rapor oluşturulamadı. "
        if errors:
            error_message += "Hatalar: " + "; ".join(errors[:3])
        return False, error_message
    
    # Raporları birleştir veya ayrı gönder
    if automated_email.merge_reports:
        # Tek dosya, birden fazla sheet
        excel_content = _merge_excel_reports(reports_list)
        attachment_filename = f"Raporlar_{today.strftime('%Y%m%d')}.xlsx"
        attachments = [(attachment_filename, excel_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')]
    else:
        # Her rapor ayrı dosya
        attachments = []
        for filename, excel_bytes in reports_list:
            attachments.append((filename, excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    
    # Mail gönder
    try:
        from django.conf import settings
        from django.core import mail
        
        # Email ayarlarını veritabanından yükle (TENANT-SPECIFIC)
        tenant = automated_email.tenant
        _update_email_settings_from_db_for_tenant(tenant)
        
        to_emails = [e.strip() for e in automated_email.to_email.split(',') if e.strip()]
        cc_emails = [e.strip() for e in automated_email.cc_email.split(',') if e.strip()] if automated_email.cc_email else []
        
        email_host = getattr(settings, 'EMAIL_HOST', '')
        email_port = getattr(settings, 'EMAIL_PORT', '')
        email_user = getattr(settings, 'EMAIL_HOST_USER', '')
        
        print(f"[DEBUG] Sending email to: {to_emails}")
        print(f"[DEBUG] Email backend: {getattr(settings, 'EMAIL_BACKEND', 'Not set')}")
        print(f"[DEBUG] Email host: {email_host}")
        print(f"[DEBUG] Email port: {email_port}")
        print(f"[DEBUG] Email user: {email_user}")
        
        if not to_emails:
            return False, "Geçerli e-posta adresi bulunamadı"
        
        # SMTP ayar kontrolü
        if not email_host:
            return False, "SMTP sunucu adresi bulunamadı. Lütfen ayarlardan 'SMTP Sunucu' alanını doldurun (örn: smtp-mail.outlook.com, smtp.gmail.com)"
        
        # Email adresi gibi görünen SMTP sunucu kontrolü
        if '@' in email_host:
            return False, f"❌ HATA: SMTP Sunucu alanına email adresi girilmiş! '{email_host}' bir email adresidir, SMTP sunucu adresi değildir. Lütfen ayarlardan 'SMTP Sunucu' alanını düzeltin: Hotmail için 'smtp-mail.outlook.com', Gmail için 'smtp.gmail.com' girin."
        
        # From email ayarla - Sistemde belirtilen "Gönderen E-posta" ayarını kullan
        default_from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None)
        
        # "Gönderen E-posta" ayarı varsa, onu kullan (kullanıcının belirttiği email)
        if default_from_email and default_from_email.strip():
            from_email = default_from_email.strip()
        else:
            # "Gönderen E-posta" ayarı yoksa veya boşsa, EMAIL_HOST_USER kullan
            from_email = email_user
        
        # Outlook/Hotmail için özel kontrol: SMTP server bazı durumlarda from_email'in 
        # EMAIL_HOST_USER ile aynı olmasını gerektirebilir, ama kullanıcı farklı belirtmişse
        # kullanıcının tercihini kullanmayı deneyelim
        # (Eğer hata alırsa, kullanıcı EMAIL_HOST_USER ile aynı yapmalı)
        
        print(f"[DEBUG] Using from_email: {from_email} (EMAIL_HOST_USER: {email_user}, DEFAULT_FROM_EMAIL: {default_from_email})")
        
        email = EmailMessage(
            subject=automated_email.subject,
            body=automated_email.body,
            from_email=from_email,
            to=to_emails,
            cc=cc_emails if cc_emails else None,
        )
        
        print(f"[DEBUG] Attaching {len(attachments)} files")
        for filename, content, content_type in attachments:
            email.attach(filename, content, content_type)
            print(f"[DEBUG] Attached: {filename} ({len(content)} bytes)")
        
        # Email backend kontrolü
        email_backend = getattr(settings, 'EMAIL_BACKEND', '')
        if 'console' in email_backend:
            print(f"[WARNING] EMAIL_BACKEND console modunda - Mail gerçekten gönderilmiyor, sadece console'a yazdırılıyor!")
        
        email.send()
        
        print(f"[DEBUG] Email.send() called successfully")
        
        # last_sent_at güncelle
        automated_email.last_sent_at = now
        automated_email.save(update_fields=['last_sent_at'])
        
        backend_info = ""
        if 'console' in email_backend:
            backend_info = " (Console modu - mail gerçekten gönderilmedi, terminal/console'a yazdırıldı)"
        
        return True, f"Mail başarıyla gönderildi ({len(attachments)} dosya){backend_info}"
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"[ERROR] Email send exception: {error_trace}")
        
        # Daha açıklayıcı hata mesajı
        error_msg = str(e)
        email_host = getattr(settings, 'EMAIL_HOST', '')
        
        if 'getaddrinfo' in error_msg or '11003' in error_msg:
            if '@' in email_host:
                return False, f"❌ SMTP Sunucu hatası: '{email_host}' bir email adresidir, sunucu adresi değildir! Lütfen ayarlardan 'SMTP Sunucu' alanını düzeltin. Hotmail için: smtp-mail.outlook.com, Gmail için: smtp.gmail.com"
            else:
                return False, f"❌ SMTP sunucu adresi çözümlenemedi: '{email_host}'. Lütfen ayarlardan 'SMTP Sunucu' alanını kontrol edin. Doğru sunucu adresi: Hotmail için 'smtp-mail.outlook.com', Gmail için 'smtp.gmail.com'"
        
        # Hotmail/Outlook authentication hatası için özel mesaj
        if '535' in error_msg and ('authentication' in error_msg.lower() or 'basic authentication is disabled' in error_msg.lower() or 'basic authentication' in error_msg.lower()):
            email_user = getattr(settings, 'EMAIL_HOST_USER', '')
            email_pass = getattr(settings, 'EMAIL_HOST_PASSWORD', '') or ''
            email_pass_length = len(email_pass)
            
            # Şifre son 4 karakterini göster (debug için)
            pass_preview = email_pass[-4:] if email_pass_length >= 4 else '***'
            
            detailed_msg = f"❌ Kimlik doğrulama hatası (535): SMTP sunucusu giriş bilgilerinizi kabul etmiyor.\n\n"
            detailed_msg += f"Kullanılan ayarlar:\n"
            detailed_msg += f"• SMTP Sunucu: {email_host}\n"
            detailed_msg += f"• SMTP Port: {getattr(settings, 'EMAIL_PORT', 'Not set')}\n"
            detailed_msg += f"• TLS: {getattr(settings, 'EMAIL_USE_TLS', 'Not set')}\n"
            detailed_msg += f"• SMTP Kullanıcı: {email_user}\n"
            detailed_msg += f"• Şifre uzunluğu: {email_pass_length} karakter\n"
            detailed_msg += f"• Şifre son 4 karakter: ...{pass_preview}\n\n"
            
            detailed_msg += f"Çözüm önerileri:\n"
            detailed_msg += f"1. 🔄 YENİ bir App Password oluşturun (eski şifre geçersiz olabilir)\n"
            detailed_msg += f"   → https://account.microsoft.com/security/app-passwords\n"
            detailed_msg += f"2. ✅ Yeni App Password'ü 'SMTP Şifresi' alanına girin\n"
            detailed_msg += f"3. ✅ Normal şifrenizi DEĞİL, App Password'ü kullanın\n"
            detailed_msg += f"4. 🔄 SMTP Sunucu'yu 'smtp.office365.com' deneyin (smtp-mail.outlook.com yerine)\n"
            detailed_msg += f"5. ✅ Ayarları kaydedin ve tekrar deneyin\n\n"
            
            detailed_msg += f"\n⚠️ KRİTİK SORUN: 'basic authentication is disabled' hatası alıyorsanız, Microsoft hesabınız Modern Authentication modunda olabilir.\n"
            detailed_msg += f"Bu durumda App Password çalışmaz. Çözüm seçenekleri:\n"
            detailed_msg += f"1. 📧 Gmail kullanın (App Password ile çalışır) - ÖNERİLEN\n"
            detailed_msg += f"   → SMTP: smtp.gmail.com, Port: 587, TLS: Açık\n"
            detailed_msg += f"2. 📧 Başka bir email servis sağlayıcısı kullanın\n"
            detailed_msg += f"3. 🔐 OAuth2 implementasyonu gerekiyor (geliştirici desteği gerekir)\n\n"
            detailed_msg += f"💡 ÖNERİ: Gmail hesabı oluşturup App Password ile kullanın. Hotmail/Outlook hesabı Modern Auth modunda olduğu için SMTP Basic Auth desteklemiyor."
            
            return False, detailed_msg
        
        return False, f"Mail gönderme hatası: {error_msg}"

@login_required
def automated_email_send_now(request, pk):
    """Otomatik maili manuel olarak şimdi gönder"""
    automated_email = get_object_or_404(AutomatedEmail, pk=pk)
    
    is_root = is_root_admin(request.user)
    if not is_root and automated_email.tenant != _resolve_tenant_for_automated_email(request):
        messages.error(request, "Bu otomatik maili gönderme yetkiniz yok.")
        return redirect("automated_email_list")
    
    if request.method == "POST":
        try:
            import traceback
            print(f"[DEBUG] Starting mail send for: {automated_email.subject}")
            print(f"[DEBUG] Selected reports: {automated_email.selected_reports}")
            
            success, message = _send_automated_email(automated_email, force=True)
            
            if success:
                messages.success(request, message)
                print(f"[DEBUG] Mail sent successfully: {message}")
            else:
                messages.error(request, f"Mail gönderilemedi: {message}")
                print(f"[DEBUG] Mail send failed: {message}")
        except Exception as e:
            error_trace = traceback.format_exc()
            error_msg = f"Mail gönderilirken hata oluştu: {str(e)}"
            print(f"[ERROR] Exception in send_now: {error_trace}")
            messages.error(request, error_msg)
    
    return redirect("automated_email_list")

@login_required
def automated_email_delete(request, pk):
    automated_email = get_object_or_404(AutomatedEmail, pk=pk)
    
    is_root = is_root_admin(request.user)
    if not is_root and automated_email.tenant != _resolve_tenant_for_automated_email(request):
        messages.error(request, "Bu otomatik maili silme yetkiniz yok.")
        return redirect("automated_email_list")
    
    if request.method == "POST":
        automated_email.delete()
        messages.success(request, "Otomatik mail başarıyla silindi.")
        return redirect("automated_email_list")
    
    return redirect("automated_email_list")